
import streamlit as st
import pandas as pd
import json
from io import BytesIO
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side
import re

# ------------------------------
# Alias dictionaries per country
# ------------------------------
allowed_fields_serbia = {
    "(AOP001)Fixed assets": ["Fixed assets"],
    "(AOP002)Intangible assets": ["I. Intangible assets", "Intangible assets", "Intangible fixed assets", "Total Intangible assets"],
    "(AOP004)Concessions,patents,licenses and similar rights and other intangible assets":["Concessions,industrial rights,licences","Concessions, patents, licenses, trade marks etc.", "Service marks,software and other intangible assets","Concessions,patents,licenses and similar rights and other intangible assets","Concessions, patents, licenses, trademarks, service marks, software and other intangible assets"],
    "(AOP005)Goodwill":["Goodwill"],
    "(AOP006)Advances for intangible assets and intangible assets in preparation":["Advances for intangible assets and intangible assets in preparation"],
    "(AOP008)Other intangible fixed assets":["Other intangible fixed assets","Other Intangible Assets"],
    "(AOP009)Tangible fixed assets": ["II. Tangible assets","Tangible fixed assets", "Total tangible assets" , "Property,plants,equipment and biological assets", "Immovables,plants and equipment","Immovables, plants and equipment","Tangible assets","Property, plants, equipment and biological assets"],
    "(AOP010)Land and buildings":["Land and buildings","Property, plant and equipment","Land & Buildings"],
    "(AOP011)Land":["Land","Land and buildings"],
    "(AOP012)Buildings":["Buildings","Land and buildings"],
    "(AOP013)Machinery and equipment": ["Machinery and equipment", "Plant and Machinery", "Technical equipment and machinery", "Property,plant and equipment","Plant and equipment","Plant & Machinery"],
    "(AOP014)Other equipment, furniture, fittings, tools, fixtures, vehicles": ["Other equipment, furniture, fittings, tools, fixtures, vehicles"],
    "(AOP016)Biological property":["Biological property","Biological assets"],
    "(AOP017)Advance payments for tangible assets": ["Advance payments for tangible assets", "Advances in property,plant,equipment and biological assets and property,biological asset in preparation","Advances in property, plant, equipment and biological assets and property, plant, equipment and biological assets in preparation"],
    "(AOP018)Tangible assets in progress": ["Tangible assets in progress", "Payments and fixed goods under construction","Immovables, plant  and equipment under construction"],
    "(AOP019)Other tangible assets": ["Other tangible assets","Other tangible fixed assets", "Other equipment" ,"Other Immovables,plants and equipment","Other immovables, plant and equipment and investment in third-party immovables, plant and equipment","Other equipment, furniture, fittings, tools, fixtures, vehicles","Other unspecified material fixed assets"],
    "(AOP020)Investments in real estate": ["Investments in real estate","Investment property measured at the cost model"],
    "(AOP021)Financial fixed assets": ["Financial fixed assets", "Long term investments","Long term financial investments and long term receivables Financial fixed assets","LONG-TERM FINANCIAL INVESTMENTS AND LONG-TERM RECEIVABLES","Financial assets"],
    "(AOP024)Loans to Group":["Loans to Group","Long-term loans to parent and subsidiary legal entities"],
    "(AOP025)Long term loans":["Long term loans","Long-term loans to legal entities with equity participation (excluding subsidiaries)"],
    "(AOP031)Long-term receivables": ["Long-term receivables", "Other long term investments and receivables","Long-term operating receivables"],
    "(AOP035)Deferred tax assets": ["DEFERRED TAX ASSETS", "Deferred tax assets"],
    "(AOP030)Other financial fixed assets":["Other financial fixed assets"],
    "(AOP034)Other long term receivables":["Other long term receivables","Other long-term receivables"],
    "(AOP036)Short-term assets": ["Short term assets", "Current assets", "Total Current assets", "Short term assets"],
    "(AOP037)Inventory": ["Inventory", "Total inventories", "Inventories"],
    "(AOP038)Raw materials, consumabeles and supplies":["Raw materials, consumabeles and supplies", "Inventory-raw materials, consumabeles and supplies","Inventory - raw materials, consumables, working inventory", "Inventory of materials(fabrication material, spare parts, small inventory and car tires)","Inventory of materials (fabrication material, spare parts, small inventory and car tires"],
    "(AOP039)Inventory of materials":["Inventory of materials(fabricationmaterial, spare parts)"],
    "(AOP040)Work in progress":["Work in progress"],
    "(AOP041)Finished products and goods":["Finished products and goods", "Finished products/merchandise", "Finished goods", "Goods"],
    "(AOP042)Traiding Goods":["Traiding Goods","Inventory-Trading Goods","Inventory - Trading Goods"],
    "(AOP048)Prepayments":["Prepayments","Payments in advance for stock"],
    "(AOP045)Short-term receivables": ["Short-term receivables"],
    "(AOP046)Short-term intercompany receivables": ["Short-term intercompany receivables", "Receivables from parent companies and subsidiaries", "Receivables from affiliated companies", "Intercompany Receivables","Receivables from foreign parent companies, subsidiaries and other associated companies","Group Receivables"],
    "(AOP047)Short-term trade receivables": ["Short-term trade receivables", "Receivables from buyers", "Trade debtor" , "Trade receivables","Trade debtors"],
    "(AOP050)Short-term receivables from employees": ["Short-term receivables from employees"],
    "(AOP049)Receivables from the state and other institutions": ["Prepaid corporate income tax","Receivables from the state and other institutions"],
    "(AOP051)Other short-term receivables": ["Other short-term receivables","Other short term receivables", "Other receivables", "Miscellaneous receivables"],
    "(AOP052)Short-term financial assets": ["SHORT TERM FINANCIAL ASSETS","short-term financial assets", "Financial investment", "Short term financial assets","Short-term financial investments","Short-term investments"],
    "(AOP058)Other short-term financial investments":["other short-term financial investments","Other short term financial assets"],
    "(AOP059_060)Cash": ["Cash","Cash and cash equivalents", "Cash assets", "Bank balance cheques and cash on hand", "Cash and cash equivalent", "Cash","Bank balance, cheques and cash on hand","Liquid assets"],
    "(AOP062)Prepaid expenses": ["Prepaid expenses","ACCRUALS"],
    "(AOP063)Total assets": ["TOTAL ASSETS", "Total assets", "Balance sheet total"],
    "(AOP064)Off balance sheet items": ["Off balance sheet items"],
    "(AOP065)Equity capital": ["Equity capital", "Owners equity"],
    "(AOP066)Subscribed and paid capital": ["Subscribed and paid capital", "Basic capital", "Shareholders equity","Called capital", "Share capital", "Called and share capital","Subscribed capital","Called up share capital (issued capital stock)"],
    "(AOP067)Emission premium":["Emission premium","Share premium"],
    "(AOP068)Own shares":["Own shares","Called up share capital","Called up share capital (issued capital stock)","Share capital","Subscribed capital unpaid"],
    "(AOP069)Subscribed unpaid capital":["Subscribed unpaid capital","Unpaid issued capital","Subscribed capital unpaid"],
    "(AOP071)Capital reserves": ["CAPITAL RESERVES", "Capital reserves", "Reserves", "Revenue","Revenue Reserves","Capital reserve"],
    "(AOP070)Revaluation reserves": ["Revaluation reserves","POSITIVE REVALUATION RESERVES AND UNREALIZED PROFIT FROM FINANCIAL ASSETS AND OTHER ELEMENTS OF OTHER COMPREHENSIVE INCOME"],
    "(AOP072)Legal reserves":["Legal reserves"],
    "(AOP074)Other reserves":["Other reserves"],
    "(AOP075+/AOP076-)Profit or loss carried forward": ["Profit or loss carried forward","Retained profit (earnings) of the year (net profits)","Accumulated retained earnings from previous periods","Retained profit", "Retained earnings brought forward", "Net retained profit/net acumalted loss", "Net profit or loss from previos years","Equity, net retained profits/net accumulated losses (balance sheet)","Net profit or loss from previous periods","Accumulated profit reserve","Retained earnings or loss"],
    "(AOP077+/AOP078-)Net profit or loss for the year": ["Net profit or loss for the year","Current period profit", "Loss from previos years", "Net profit for the period" , "Profit of the year","Net profit or loss for the year", "Net profit for the period","Retained earnings for the current year"],
    "(AOP081)Liabilities": ["Liabilities", "Total liabilities","Total debt","Total debts"],
    "(AOP082)Provision for risks and charges":["Provision for risks and charges","Provisions for risks and charges"],
    "(AOP083)Provisions for pensions and similar obligations":["Provisions for pensions and similar obligations"],
    "(AOP084)Other provisions":["Other provisions"],
    "(AOP085)Long-term liabilities": ["Long-term liabilities", "Long term liabilities", "Total Long term liabilities","Long-term financial and operating liabilities","Long-term provisions and long-term liabilities"],
    "(AOP086)Long-term liabilities to affiliates": ["Long-term liabilities to affiliates", "Group payables due after 1 year", "Long term liabilities to affiliates"],
    "(AOP087)Trade liabilities":["Trade liabilities"],
    "(AOP090)Long-term liabilities for loans": ["Long-term liabilities for loans"],
    "(AOP092)Other loans/finance due after 1 year":["Other loans/finance due after 1 year","Long term liabilities payable to financial institutions"],
    "(AOP093)Other long-term liabilities": ["Other long-term liabilities", "Other long term liabilities"],
    "(AOP095)Short-term liabilities": ["Short-term liabilities","IV. SHORT-TERM LIABILITIES","Short- term liabilities","Short term liabilities","SHORT- TERM LIABILITIES", "TOTAL CURRENT LIABILITIES", "Total current liabilities","Short- term liabilities and short term provisions"," Short- term financial and operating liabilities"],
    "(AOP096)Short-term liabilities to affiliates": ["Short-term liabilities to affiliates","Liabilities of disposal groups", " Trade payables-foreing parent company, subsidiares and other associated companies", " Liabilities to affiliated companies","Group payables"],
    "(AOP098)Prepayments, deposits and gurantee":["Prepayments, deposits and gurantee","Liabilities for securities","Prepayments, deposits and guarantees"],
    "(AOP103)Liabilities for loans, deposits, etc. to companies within the group": ["Liabilities for loans, deposits, etc. to companies within the group"],
    "(AOP097)Short-term trade creditors": ["Short-term trade creditors", "Liabilities to suppliers","Trade Payables","Trade creditors(acounts payable)", " Trade payables","Commitments towards suppliers","Trade payable"],
    "(AOP100)Short-term liabilities to employees": ["Short-term liabilities to employees","Liabilities towards employees"],
    "(AOP099)Short-term liabilities for taxes, contributions and other fees": ["Short-term liabilities for taxes, contributions and other fees"],
    "(AOP101)Liability for income tax":["Liability for income tax","S-term Liabilities for tax,contributions and other fees"],
    "(AOP104)Obligations for taken loans and credits":["Obligations for taken loans and credits","Short Term Bank Debt","Short liabilities for loans","Short term loans","Short term financial liabilities","Obligation for taken loans","BANK LOANS AND OVERDRAFT","Liabilities to bank","Loans liabilities from credit institutions","Short-term liabilities for loans","Short-term financial liabilities","S-term Liabilities payable to financial institutions","Banks loans and overdraft","Other Loans/Finance"],
    "(AOP108)Other short-term liabilities": ["Other short term liabilities", "Other operating and liabilities and other short term liabilities", "Other short-term liabilities ",  "Other  liabilities  including accruals",  "Other  liabilities","Other short-term liabilities","Miscellaneous Liabilities"],
    "(AOP109)Accruals and deferred income": ["Accruals and deferred income"],
    "(AOP111)Total liabilities and funds": ["Total liabilities and funds", "TOTAL EQUITY AND LIABILITIES "," TOTAL LIABILITIES","BALANCE SHEET TOTAL","TOTAL LIABILITIES","Balance sheet total"],
    "(AOP112)Off balance sheet items": ["Off balance sheet items"],
    "(AOP201)Turnover, sales revenue": ["Turnover, sales revenue","Revenues from contracts with customers","Operating income","Turnover","Total income"],
    "(AOP202)Revenues from sales": ["Revenues from sales","Income from sales","Income from sales (outside group)","Income from goods sold","Income from sales of goods","Turnover","Operating income","Net sales revenue"],
    "(AOP203)Other income(other revenues)": ["Other income(other revenues)","Other operating income (outside the group)","Other operating income","OTHER OPERATING REVENUE","Other income and profits","Income from financial transactions (financial income)","Other income"],
    "(AOP206)Own work capitalized": ["Own work capitalized"],
    "(AOP207)Operating expenses": ["Operating expenses","Operating Costs"],
    "(AOP208)Material costs": ["Material costs","Cost of raw materials and consumables","RAW MATERIAL COSTS, FUEL AND ENERGY COSTS"],
    "(AOP209)Cost of goods sold": ["Cost of goods sold","The purchase value of the goods sold","Cost of materials (type of expenditure format)","Cost of goods sold and the cost of materials"],
    "(AOP213)Staff costs": ["Staff costs (employee costs)","Wages and salaries","Wages & Salaries","Short-term financial and operating liabilities","Personnel type expenses","Wages expenses, wage compensation and other personal expenses"],
    "(AOP218)Depreciation on fixed assets": ["Depreciation on fixed assets","Depreciation","DepreciationDepreciation and provisions","Depreciation and amortization"],
    "(AOP222)Other operating expenses": ["Other operating expenses","Other expenses","Intangible costs","Other expenses and losses"],
    "(AOP223)Income from financial transactions": ["Income from financial transactions (financial income)","III. FINANCIAL INCOME","Financial income"],
    "(AOP234)Financial costs": ["Financial costs"],
    "(AOP250+/AOP251-)Profit or loss before taxation": ["Profit or loss before taxation","Profit before taxation","Loss before taxation","Profit before taxation/Loss before taxation","Profit Before Tax"],
    "(AOP252)Profit tax": ["Profit tax","Income tax","Taxes,duties and similar expenses","Taxes","Tax","Tax charge","Tax expense of the period"],
    "(AOP255+/AOP256-)Profit or loss after taxation": ["Profit or loss after taxation","Profit after taxation","Loss after taxation","TOTAL RESULT"],
}
allowed_fields_netherlands = {
    "(AOP001)Fixed assets": ["Fixed assets"],
    "(AOP002)Intangible assets": ["Total Intangible assets"],
    "(AOP009)Tangible fixed assets": ["II. Tangible assets","Tangible fixed assets", "Total tangible assets" , "Property,plants,equipment and biological assets", "Immovables,plants and equipment","Immovables, plants and equipment","Tangible assets","Property, plants, equipment and biological assets"],
    "(AOP036)Short-term assets": ["Short term assets", "Current assets", "Total Current assets", "Short term assets"],
    "(AOP037)Inventory": ["Inventory", "Total inventories", "Inventories"],
    "(AOP045)Short-term receivables": ["Short-term receivables"],
    "(AOP047)Short-term trade receivables": ["Short-term trade receivables", "Receivables from buyers", "Trade debtor" , "Trade receivables","Trade debtors"],
    "(AOP051)Other short-term receivables": ["Other short-term receivables","Other short term receivables", "Other receivables", "Miscellaneous receivables"],
    "(AOP059)Cash": ["Cash","Cash and cash equivalents", "Cash assets", "Bank balance cheques and cash on hand", "Cash and cash equivalent", "Cash","Bank balance, cheques and cash on hand","Liquid assets"],
    "(AOP063)Total assets": ["TOTAL ASSETS", "Total assets", "Balance sheet total"],
    "(AOP065)Equity capital": ["Equity capital", "Owners equity"],
    "(AOP071)Capital reserves": ["CAPITAL RESERVES", "Capital reserves", "Reserves", "Revenue","Revenue Reserves","Capital reserve"],
    "(AOP075)Profit or loss carried forward": ["Profit or loss carried forward","Retained profit (earnings) of the year (net profits)","Accumulated retained earnings from previous periods","Retained profit", "Retained earnings brought forward", "Net retained profit/net acumalted loss", "Net profit or loss from previos years","Equity, net retained profits/net accumulated losses (balance sheet)","Net profit or loss from previous periods","Accumulated profit reserve","Retained earnings or loss"],
    "(AOP081)Liabilities": ["Liabilities", "Total liabilities","Total debt","Total debts"],
    "(AOP085)Long-term liabilities": ["Long-term liabilities", "Long term liabilities", "Total Long term liabilities","Long-term financial and operating liabilities","Long-term provisions and long-term liabilities"],
    "(AOP093)Other long-term liabilities": ["Other long-term liabilities", "Other long term liabilities"],
    "(AOP095)Short-term liabilities": ["Short-term liabilities","IV. SHORT-TERM LIABILITIES","Short- term liabilities","Short term liabilities","SHORT- TERM LIABILITIES", "TOTAL CURRENT LIABILITIES", "Total current liabilities","Short- term liabilities and short term provisions"," Short- term financial and operating liabilities"],
    "(AOP096)Short-term liabilities to affiliates": ["Short-term liabilities to affiliates","Liabilities of disposal groups", " Trade payables-foreing parent company, subsidiares and other associated companies", " Liabilities to affiliated companies","Group payables"],
    "(AOP097)Short-term trade creditors": ["Short-term trade creditors", "Liabilities to suppliers","Trade Payables","Trade creditors(acounts payable)", " Trade payables","Commitments towards suppliers","Trade payable"],
    "(AOP108)Other short-term liabilities": ["Other short term liabilities", "Other operating and liabilities and other short term liabilities", "Other short-term liabilities ",  "Other  liabilities  including accruals",  "Other  liabilities","Other short-term liabilities","Miscellaneous Liabilities"],
    "(AOP111)Total liabilities and funds": ["Total liabilities and funds", "TOTAL EQUITY AND LIABILITIES "," TOTAL LIABILITIES","BALANCE SHEET TOTAL","TOTAL LIABILITIES","Balance sheet total"],
    "(AOP201)Turnover, sales revenue": ["Turnover, sales revenue","Revenues from contracts with customers","Operating income","Turnover","Total income"],
    "(AOP207)Operating expenses": ["Operating expenses","Operating Costs"],
    "(AOP234)Financial costs": ["Financial costs"],
    "(AOP250)Profit or loss before taxation": ["Profit or loss before taxation","Profit before taxation","Loss before taxation","Profit before taxation/Loss before taxation","Profit Before Tax"],
    "(AOP252)Profit tax": ["Profit tax","Income tax","Taxes,duties and similar expenses","Taxes","Tax","Tax charge","Tax expense of the period"],
    "(AOP255)Profit or loss after taxation": ["Profit or loss after taxation","Profit after taxation","Loss after taxation","TOTAL RESULT"],
}
allowed_fields_france = {
    "(AOP001)Fixed assets": ["Fixed assets"],
    "(AOP002)Intangible assets": ["I. Intangible assets", "Intangible assets", "Intangible fixed assets", "Total Intangible assets"],
    "(AOP004)Concessions,patents,licenses and similar rights and other intangible assets":["Concessions,industrial rights,licences","Concessions, patents, licenses, trade marks etc.", "Service marks,software and other intangible assets","Concessions,patents,licenses and similar rights and other intangible assets","Concessions, patents, licenses, trademarks, service marks, software and other intangible assets"],
    "(AOP005)Goodwill":["Goodwill"],
    "(AOP006)Advances for intangible assets and intangible assets in preparation":["Advances for intangible assets and intangible assets in preparation"],
    "(AOP008)Other intangible fixed assets":["Other intangible fixed assets","Other Intangible Assets"],
    "(AOP009)Tangible fixed assets": ["II. Tangible assets","Tangible fixed assets", "Total tangible assets" , "Property,plants,equipment and biological assets", "Immovables,plants and equipment","Immovables, plants and equipment","Tangible assets","Property, plants, equipment and biological assets"],
    "(AOP010)Land and buildings":["Land and buildings","Property, plant and equipment","Land & Buildings"],
    "(AOP011)Land":["Land","Land and buildings"],
    "(AOP012)Buildings":["Buildings","Land and buildings"],
    "(AOP013)Machinery and equipment": ["Machinery and equipment", "Plant and Machinery", "Technical equipment and machinery", "Property,plant and equipment","Plant and equipment","Plant & Machinery"],
    "(AOP014)Other equipment, furniture, fittings, tools, fixtures, vehicles": ["Other equipment, furniture, fittings, tools, fixtures, vehicles"],
    "(AOP016)Biological property":["Biological property","Biological assets"],
    "(AOP017)Advance payments for tangible assets": ["Advance payments for tangible assets", "Advances in property,plant,equipment and biological assets and property,biological asset in preparation","Advances in property, plant, equipment and biological assets and property, plant, equipment and biological assets in preparation"],
    "(AOP018)Tangible assets in progress": ["Tangible assets in progress", "Payments and fixed goods under construction","Immovables, plant  and equipment under construction"],
    "(AOP019)Other tangible assets": ["Other tangible assets","Other tangible fixed assets", "Other equipment" ,"Other Immovables,plants and equipment","Other immovables, plant and equipment and investment in third-party immovables, plant and equipment","Other equipment, furniture, fittings, tools, fixtures, vehicles","Other unspecified material fixed assets"],
    "(AOP020)Investments in real estate": ["Investments in real estate","Investment property measured at the cost model"],
    "(AOP021)Financial fixed assets": ["Financial fixed assets", "Long term investments","Long term financial investments and long term receivables Financial fixed assets","LONG-TERM FINANCIAL INVESTMENTS AND LONG-TERM RECEIVABLES","Financial assets"],
    "(AOP024)Loans to Group":["Loans to Group","Long-term loans to parent and subsidiary legal entities"],
    "(AOP025)Long term loans":["Long term loans","Long-term loans to legal entities with equity participation (excluding subsidiaries)"],
    "(AOP031)Long-term receivables": ["Long-term receivables", "Other long term investments and receivables","Long-term operating receivables"],
    "(AOP035)Deferred tax assets": ["DEFERRED TAX ASSETS", "Deferred tax assets"],
    "(AOP030)Other financial fixed assets":["Other financial fixed assets"],
    "(AOP034)Other long term receivables":["Other long term receivables","Other long-term receivables"],
    "(AOP036)Short-term assets": ["Short term assets", "Current assets", "Total Current assets", "Short term assets"],
    "(AOP037)Inventory": ["Inventory", "Total inventories", "Inventories"],
    "(AOP038)Raw materials, consumabeles and supplies":["Raw materials, consumabeles and supplies", "Inventory-raw materials, consumabeles and supplies","Inventory - raw materials, consumables, working inventory", "Inventory of materials(fabrication material, spare parts, small inventory and car tires)","Inventory of materials (fabrication material, spare parts, small inventory and car tires"],
    "(AOP039)Inventory of materials":["Inventory of materials(fabricationmaterial, spare parts)"],
    "(AOP040)Work in progress":["Work in progress"],
    "(AOP041)Finished products and goods":["Finished products and goods", "Finished products/merchandise", "Finished goods", "Goods"],
    "(AOP042)Traiding Goods":["Traiding Goods","Inventory-Trading Goods","Inventory - Trading Goods"],
    "(AOP048)Prepayments":["Prepayments","Payments in advance for stock"],
    "(AOP045)Short-term receivables": ["Short-term receivables"],
    "(AOP046)Short-term intercompany receivables": ["Short-term intercompany receivables", "Receivables from parent companies and subsidiaries", "Receivables from affiliated companies", "Intercompany Receivables","Receivables from foreign parent companies, subsidiaries and other associated companies","Group Receivables"],
    "(AOP047)Short-term trade receivables": ["Short-term trade receivables", "Receivables from buyers", "Trade debtor" , "Trade receivables","Trade debtors"],
    "(AOP050)Short-term receivables from employees": ["Short-term receivables from employees"],
    "(AOP049)Receivables from the state and other institutions": ["Prepaid corporate income tax","Receivables from the state and other institutions"],
    "(AOP051)Other short-term receivables": ["Other short-term receivables","Other short term receivables", "Other receivables", "Miscellaneous receivables"],
    "(AOP052)Short-term financial assets": ["SHORT TERM FINANCIAL ASSETS","short-term financial assets", "Financial investment", "Short term financial assets","Short-term financial investments","Short-term investments"],
    "(AOP058)Other short-term financial investments":["other short-term financial investments","Other short term financial assets"],
    "(AOP059_060)Cash": ["Cash","Cash and cash equivalents", "Cash assets", "Bank balance cheques and cash on hand", "Cash and cash equivalent", "Cash","Bank balance, cheques and cash on hand","Liquid assets"],
    "(AOP062)Prepaid expenses": ["Prepaid expenses","ACCRUALS"],
    "(AOP063)Total assets": ["TOTAL ASSETS", "Total assets", "Balance sheet total"],
    "(AOP064)Off balance sheet items": ["Off balance sheet items"],
    "(AOP065)Equity capital": ["Equity capital", "Owners equity"],
    "(AOP066)Subscribed and paid capital": ["Subscribed and paid capital", "Basic capital", "Shareholders equity","Called capital", "Share capital", "Called and share capital","Subscribed capital","Called up share capital (issued capital stock)"],
    "(AOP067)Emission premium":["Emission premium","Share premium"],
    "(AOP068)Own shares":["Own shares","Called up share capital","Called up share capital (issued capital stock)","Share capital","Subscribed capital unpaid"],
    "(AOP069)Subscribed unpaid capital":["Subscribed unpaid capital","Unpaid issued capital","Subscribed capital unpaid"],
    "(AOP071)Capital reserves": ["CAPITAL RESERVES", "Capital reserves", "Reserves", "Revenue","Revenue Reserves","Capital reserve"],
    "(AOP070)Revaluation reserves": ["Revaluation reserves","POSITIVE REVALUATION RESERVES AND UNREALIZED PROFIT FROM FINANCIAL ASSETS AND OTHER ELEMENTS OF OTHER COMPREHENSIVE INCOME"],
    "(AOP072)Legal reserves":["Legal reserves"],
    "(AOP074)Other reserves":["Other reserves"],
    "(AOP075+/AOP076-)Profit or loss carried forward": ["Profit or loss carried forward","Retained profit (earnings) of the year (net profits)","Accumulated retained earnings from previous periods","Retained profit", "Retained earnings brought forward", "Net retained profit/net acumalted loss", "Net profit or loss from previos years","Equity, net retained profits/net accumulated losses (balance sheet)","Net profit or loss from previous periods","Accumulated profit reserve","Retained earnings or loss"],
    "(AOP077+/AOP078-)Net profit or loss for the year": ["Net profit or loss for the year","Current period profit", "Loss from previos years", "Net profit for the period" , "Profit of the year","Net profit or loss for the year", "Net profit for the period","Retained earnings for the current year"],
    "(AOP081)Liabilities": ["Liabilities", "Total liabilities","Total debt","Total debts"],
    "(AOP082)Provision for risks and charges":["Provision for risks and charges","Provisions for risks and charges"],
    "(AOP083)Provisions for pensions and similar obligations":["Provisions for pensions and similar obligations"],
    "(AOP084)Other provisions":["Other provisions"],
    "(AOP085)Long-term liabilities": ["Long-term liabilities", "Long term liabilities", "Total Long term liabilities","Long-term financial and operating liabilities","Long-term provisions and long-term liabilities"],
    "(AOP086)Long-term liabilities to affiliates": ["Long-term liabilities to affiliates", "Group payables due after 1 year", "Long term liabilities to affiliates"],
    "(AOP087)Trade liabilities":["Trade liabilities"],
    "(AOP090)Long-term liabilities for loans": ["Long-term liabilities for loans"],
    "(AOP092)Other loans/finance due after 1 year":["Other loans/finance due after 1 year","Long term liabilities payable to financial institutions"],
    "(AOP093)Other long-term liabilities": ["Other long-term liabilities", "Other long term liabilities"],
    "(AOP095)Short-term liabilities": ["Short-term liabilities","IV. SHORT-TERM LIABILITIES","Short- term liabilities","Short term liabilities","SHORT- TERM LIABILITIES", "TOTAL CURRENT LIABILITIES", "Total current liabilities","Short- term liabilities and short term provisions"," Short- term financial and operating liabilities"],
    "(AOP096)Short-term liabilities to affiliates": ["Short-term liabilities to affiliates","Liabilities of disposal groups", " Trade payables-foreing parent company, subsidiares and other associated companies", " Liabilities to affiliated companies","Group payables"],
    "(AOP098)Prepayments, deposits and gurantee":["Prepayments, deposits and gurantee","Liabilities for securities","Prepayments, deposits and guarantees"],
    "(AOP103)Liabilities for loans, deposits, etc. to companies within the group": ["Liabilities for loans, deposits, etc. to companies within the group"],
    "(AOP097)Short-term trade creditors": ["Short-term trade creditors", "Liabilities to suppliers","Trade Payables","Trade creditors(acounts payable)", " Trade payables","Commitments towards suppliers","Trade payable"],
    "(AOP100)Short-term liabilities to employees": ["Short-term liabilities to employees","Liabilities towards employees"],
    "(AOP099)Short-term liabilities for taxes, contributions and other fees": ["Short-term liabilities for taxes, contributions and other fees"],
    "(AOP101)Liability for income tax":["Liability for income tax","S-term Liabilities for tax,contributions and other fees"],
    "(AOP104)Obligations for taken loans and credits":["Obligations for taken loans and credits","Short Term Bank Debt","Short liabilities for loans","Short term loans","Short term financial liabilities","Obligation for taken loans","BANK LOANS AND OVERDRAFT","Liabilities to bank","Loans liabilities from credit institutions","Short-term liabilities for loans","Short-term financial liabilities","S-term Liabilities payable to financial institutions","Banks loans and overdraft","Other Loans/Finance"],
    "(AOP108)Other short-term liabilities": ["Other short term liabilities", "Other operating and liabilities and other short term liabilities", "Other short-term liabilities ",  "Other  liabilities  including accruals",  "Other  liabilities","Other short-term liabilities","Miscellaneous Liabilities"],
    "(AOP109)Accruals and deferred income": ["Accruals and deferred income"],
    "(AOP111)Total liabilities and funds": ["Total liabilities and funds", "TOTAL EQUITY AND LIABILITIES "," TOTAL LIABILITIES","BALANCE SHEET TOTAL","TOTAL LIABILITIES","Balance sheet total"],
    "(AOP112)Off balance sheet items": ["Off balance sheet items"],
    "(AOP201)Turnover, sales revenue": ["Turnover, sales revenue","Revenues from contracts with customers","Operating income","Turnover","Total income"],
    "(AOP202)Revenues from sales": ["Revenues from sales","Income from sales","Income from sales (outside group)","Income from goods sold","Income from sales of goods","Turnover","Operating income","Net sales revenue"],
    "(AOP203)Other income(other revenues)": ["Other income(other revenues)","Other operating income (outside the group)","Other operating income","OTHER OPERATING REVENUE","Other income and profits","Income from financial transactions (financial income)","Other income"],
    "(AOP206)Own work capitalized": ["Own work capitalized"],
    "(AOP207)Operating expenses": ["Operating expenses","Operating Costs"],
    "(AOP208)Material costs": ["Material costs","Cost of raw materials and consumables","RAW MATERIAL COSTS, FUEL AND ENERGY COSTS"],
    "(AOP209)Cost of goods sold": ["Cost of goods sold","The purchase value of the goods sold","Cost of materials (type of expenditure format)","Cost of goods sold and the cost of materials"],
    "(AOP213)Staff costs": ["Staff costs (employee costs)","Wages and salaries","Wages & Salaries","Short-term financial and operating liabilities","Personnel type expenses","Wages expenses, wage compensation and other personal expenses"],
    "(AOP218)Depreciation on fixed assets": ["Depreciation on fixed assets","Depreciation","DepreciationDepreciation and provisions","Depreciation and amortization"],
    "(AOP222)Other operating expenses": ["Other operating expenses","Other expenses","Intangible costs","Other expenses and losses"],
    "(AOP223)Income from financial transactions": ["Income from financial transactions (financial income)","III. FINANCIAL INCOME","Financial income"],
    "(AOP234)Financial costs": ["Financial costs"],
    "(AOP250+/AOP251-)Profit or loss before taxation": ["Profit or loss before taxation","Profit before taxation","Loss before taxation","Profit before taxation/Loss before taxation","Profit Before Tax"],
    "(AOP252)Profit tax": ["Profit tax","Income tax","Taxes,duties and similar expenses","Taxes","Tax","Tax charge","Tax expense of the period"],
    "(AOP255+/AOP256-)Profit or loss after taxation": ["Profit or loss after taxation","Profit after taxation","Loss after taxation","TOTAL RESULT"],
}
allowed_fields_germany = {
    "(AOP001)Fixed assets": ["Fixed assets"],
    "(AOP002)Intangible assets": ["I. Intangible assets", "Intangible assets", "Intangible fixed assets", "Total Intangible assets"],
    "(AOP004)Concessions,patents,licenses and similar rights and other intangible assets":["Concessions,industrial rights,licences","Concessions, patents, licenses, trade marks etc.", "Service marks,software and other intangible assets","Concessions,patents,licenses and similar rights and other intangible assets","Concessions, patents, licenses, trademarks, service marks, software and other intangible assets"],
    "(AOP005)Goodwill":["Goodwill"],
    "(AOP006)Advances for intangible assets and intangible assets in preparation":["Advances for intangible assets and intangible assets in preparation"],
    "(AOP008)Other intangible fixed assets":["Other intangible fixed assets","Other Intangible Assets"],
    "(AOP009)Tangible fixed assets": ["II. Tangible assets","Tangible fixed assets", "Total tangible assets" , "Property,plants,equipment and biological assets", "Immovables,plants and equipment","Immovables, plants and equipment","Tangible assets","Property, plants, equipment and biological assets"],
    "(AOP010)Land and buildings":["Land and buildings","Property, plant and equipment","Land & Buildings"],
    "(AOP011)Land":["Land","Land and buildings"],
    "(AOP012)Buildings":["Buildings","Land and buildings"],
    "(AOP013)Machinery and equipment": ["Machinery and equipment", "Plant and Machinery", "Technical equipment and machinery", "Property,plant and equipment","Plant and equipment","Plant & Machinery"],
    "(AOP014)Other equipment, furniture, fittings, tools, fixtures, vehicles": ["Other equipment, furniture, fittings, tools, fixtures, vehicles"],
    "(AOP016)Biological property":["Biological property","Biological assets"],
    "(AOP017)Advance payments for tangible assets": ["Advance payments for tangible assets", "Advances in property,plant,equipment and biological assets and property,biological asset in preparation","Advances in property, plant, equipment and biological assets and property, plant, equipment and biological assets in preparation"],
    "(AOP018)Tangible assets in progress": ["Tangible assets in progress", "Payments and fixed goods under construction","Immovables, plant  and equipment under construction"],
    "(AOP019)Other tangible assets": ["Other tangible assets","Other tangible fixed assets", "Other equipment" ,"Other Immovables,plants and equipment","Other immovables, plant and equipment and investment in third-party immovables, plant and equipment","Other equipment, furniture, fittings, tools, fixtures, vehicles","Other unspecified material fixed assets"],
    "(AOP020)Investments in real estate": ["Investments in real estate","Investment property measured at the cost model"],
    "(AOP021)Financial fixed assets": ["Financial fixed assets", "Long term investments","Long term financial investments and long term receivables Financial fixed assets","LONG-TERM FINANCIAL INVESTMENTS AND LONG-TERM RECEIVABLES","Financial assets"],
    "(AOP024)Loans to Group":["Loans to Group","Long-term loans to parent and subsidiary legal entities"],
    "(AOP025)Long term loans":["Long term loans","Long-term loans to legal entities with equity participation (excluding subsidiaries)"],
    "(AOP031)Long-term receivables": ["Long-term receivables", "Other long term investments and receivables","Long-term operating receivables"],
    "(AOP035)Deferred tax assets": ["DEFERRED TAX ASSETS", "Deferred tax assets"],
    "(AOP030)Other financial fixed assets":["Other financial fixed assets"],
    "(AOP034)Other long term receivables":["Other long term receivables","Other long-term receivables"],
    "(AOP036)Short-term assets": ["Short term assets", "Current assets", "Total Current assets", "Short term assets"],
    "(AOP037)Inventory": ["Inventory", "Total inventories", "Inventories"],
    "(AOP038)Raw materials, consumabeles and supplies":["Raw materials, consumabeles and supplies", "Inventory-raw materials, consumabeles and supplies","Inventory - raw materials, consumables, working inventory", "Inventory of materials(fabrication material, spare parts, small inventory and car tires)","Inventory of materials (fabrication material, spare parts, small inventory and car tires"],
    "(AOP039)Inventory of materials":["Inventory of materials(fabricationmaterial, spare parts)"],
    "(AOP040)Work in progress":["Work in progress"],
    "(AOP041)Finished products and goods":["Finished products and goods", "Finished products/merchandise", "Finished goods", "Goods"],
    "(AOP042)Traiding Goods":["Traiding Goods","Inventory-Trading Goods","Inventory - Trading Goods"],
    "(AOP048)Prepayments":["Prepayments","Payments in advance for stock"],
    "(AOP045)Short-term receivables": ["Short-term receivables"],
    "(AOP046)Short-term intercompany receivables": ["Short-term intercompany receivables", "Receivables from parent companies and subsidiaries", "Receivables from affiliated companies", "Intercompany Receivables","Receivables from foreign parent companies, subsidiaries and other associated companies","Group Receivables"],
    "(AOP047)Short-term trade receivables": ["Short-term trade receivables", "Receivables from buyers", "Trade debtor" , "Trade receivables","Trade debtors"],
    "(AOP050)Short-term receivables from employees": ["Short-term receivables from employees"],
    "(AOP049)Receivables from the state and other institutions": ["Prepaid corporate income tax","Receivables from the state and other institutions"],
    "(AOP051)Other short-term receivables": ["Other short-term receivables","Other short term receivables", "Other receivables", "Miscellaneous receivables"],
    "(AOP052)Short-term financial assets": ["SHORT TERM FINANCIAL ASSETS","short-term financial assets", "Financial investment", "Short term financial assets","Short-term financial investments","Short-term investments"],
    "(AOP058)Other short-term financial investments":["other short-term financial investments","Other short term financial assets"],
    "(AOP059_060)Cash": ["Cash","Cash and cash equivalents", "Cash assets", "Bank balance cheques and cash on hand", "Cash and cash equivalent", "Cash","Bank balance, cheques and cash on hand","Liquid assets"],
    "(AOP062)Prepaid expenses": ["Prepaid expenses","ACCRUALS"],
    "(AOP063)Total assets": ["TOTAL ASSETS", "Total assets", "Balance sheet total"],
    "(AOP064)Off balance sheet items": ["Off balance sheet items"],
    "(AOP065)Equity capital": ["Equity capital", "Owners equity"],
    "(AOP066)Subscribed and paid capital": ["Subscribed and paid capital", "Basic capital", "Shareholders equity","Called capital", "Share capital", "Called and share capital","Subscribed capital","Called up share capital (issued capital stock)"],
    "(AOP067)Emission premium":["Emission premium","Share premium"],
    "(AOP068)Own shares":["Own shares","Called up share capital","Called up share capital (issued capital stock)","Share capital","Subscribed capital unpaid"],
    "(AOP069)Subscribed unpaid capital":["Subscribed unpaid capital","Unpaid issued capital","Subscribed capital unpaid"],
    "(AOP071)Capital reserves": ["CAPITAL RESERVES", "Capital reserves", "Reserves", "Revenue","Revenue Reserves","Capital reserve"],
    "(AOP070)Revaluation reserves": ["Revaluation reserves","POSITIVE REVALUATION RESERVES AND UNREALIZED PROFIT FROM FINANCIAL ASSETS AND OTHER ELEMENTS OF OTHER COMPREHENSIVE INCOME"],
    "(AOP072)Legal reserves":["Legal reserves"],
    "(AOP074)Other reserves":["Other reserves"],
    "(AOP075+/AOP076-)Profit or loss carried forward": ["Profit or loss carried forward","Retained profit (earnings) of the year (net profits)","Accumulated retained earnings from previous periods","Retained profit", "Retained earnings brought forward", "Net retained profit/net acumalted loss", "Net profit or loss from previos years","Equity, net retained profits/net accumulated losses (balance sheet)","Net profit or loss from previous periods","Accumulated profit reserve","Retained earnings or loss"],
    "(AOP077+/AOP078-)Net profit or loss for the year": ["Net profit or loss for the year","Current period profit", "Loss from previos years", "Net profit for the period" , "Profit of the year","Net profit or loss for the year", "Net profit for the period","Retained earnings for the current year"],
    "(AOP081)Liabilities": ["Liabilities", "Total liabilities","Total debt","Total debts"],
    "(AOP082)Provision for risks and charges":["Provision for risks and charges","Provisions for risks and charges"],
    "(AOP083)Provisions for pensions and similar obligations":["Provisions for pensions and similar obligations"],
    "(AOP084)Other provisions":["Other provisions"],
    "(AOP085)Long-term liabilities": ["Long-term liabilities", "Long term liabilities", "Total Long term liabilities","Long-term financial and operating liabilities","Long-term provisions and long-term liabilities"],
    "(AOP086)Long-term liabilities to affiliates": ["Long-term liabilities to affiliates", "Group payables due after 1 year", "Long term liabilities to affiliates"],
    "(AOP087)Trade liabilities":["Trade liabilities"],
    "(AOP090)Long-term liabilities for loans": ["Long-term liabilities for loans"],
    "(AOP092)Other loans/finance due after 1 year":["Other loans/finance due after 1 year","Long term liabilities payable to financial institutions"],
    "(AOP093)Other long-term liabilities": ["Other long-term liabilities", "Other long term liabilities"],
    "(AOP095)Short-term liabilities": ["Short-term liabilities","IV. SHORT-TERM LIABILITIES","Short- term liabilities","Short term liabilities","SHORT- TERM LIABILITIES", "TOTAL CURRENT LIABILITIES", "Total current liabilities","Short- term liabilities and short term provisions"," Short- term financial and operating liabilities"],
    "(AOP096)Short-term liabilities to affiliates": ["Short-term liabilities to affiliates","Liabilities of disposal groups", " Trade payables-foreing parent company, subsidiares and other associated companies", " Liabilities to affiliated companies","Group payables"],
    "(AOP098)Prepayments, deposits and gurantee":["Prepayments, deposits and gurantee","Liabilities for securities","Prepayments, deposits and guarantees"],
    "(AOP103)Liabilities for loans, deposits, etc. to companies within the group": ["Liabilities for loans, deposits, etc. to companies within the group"],
    "(AOP097)Short-term trade creditors": ["Short-term trade creditors", "Liabilities to suppliers","Trade Payables","Trade creditors(acounts payable)", " Trade payables","Commitments towards suppliers","Trade payable"],
    "(AOP100)Short-term liabilities to employees": ["Short-term liabilities to employees","Liabilities towards employees"],
    "(AOP099)Short-term liabilities for taxes, contributions and other fees": ["Short-term liabilities for taxes, contributions and other fees"],
    "(AOP101)Liability for income tax":["Liability for income tax","S-term Liabilities for tax,contributions and other fees"],
    "(AOP104)Obligations for taken loans and credits":["Obligations for taken loans and credits","Short Term Bank Debt","Short liabilities for loans","Short term loans","Short term financial liabilities","Obligation for taken loans","BANK LOANS AND OVERDRAFT","Liabilities to bank","Loans liabilities from credit institutions","Short-term liabilities for loans","Short-term financial liabilities","S-term Liabilities payable to financial institutions","Banks loans and overdraft","Other Loans/Finance"],
    "(AOP108)Other short-term liabilities": ["Other short term liabilities", "Other operating and liabilities and other short term liabilities", "Other short-term liabilities ",  "Other  liabilities  including accruals",  "Other  liabilities","Other short-term liabilities","Miscellaneous Liabilities"],
    "(AOP109)Accruals and deferred income": ["Accruals and deferred income"],
    "(AOP111)Total liabilities and funds": ["Total liabilities and funds", "TOTAL EQUITY AND LIABILITIES "," TOTAL LIABILITIES","BALANCE SHEET TOTAL","TOTAL LIABILITIES","Balance sheet total"],
    "(AOP112)Off balance sheet items": ["Off balance sheet items"],
    "(AOP201)Turnover, sales revenue": ["Turnover, sales revenue","Revenues from contracts with customers","Operating income","Turnover","Total income"],
    "(AOP202)Revenues from sales": ["Revenues from sales","Income from sales","Income from sales (outside group)","Income from goods sold","Income from sales of goods","Turnover","Operating income","Net sales revenue"],
    "(AOP203)Other income(other revenues)": ["Other income(other revenues)","Other operating income (outside the group)","Other operating income","OTHER OPERATING REVENUE","Other income and profits","Income from financial transactions (financial income)","Other income"],
    "(AOP206)Own work capitalized": ["Own work capitalized"],
    "(AOP207)Operating expenses": ["Operating expenses","Operating Costs"],
    "(AOP208)Material costs": ["Material costs","Cost of raw materials and consumables","RAW MATERIAL COSTS, FUEL AND ENERGY COSTS"],
    "(AOP209)Cost of goods sold": ["Cost of goods sold","The purchase value of the goods sold","Cost of materials (type of expenditure format)","Cost of goods sold and the cost of materials"],
    "(AOP213)Staff costs": ["Staff costs (employee costs)","Wages and salaries","Wages & Salaries","Short-term financial and operating liabilities","Personnel type expenses","Wages expenses, wage compensation and other personal expenses"],
    "(AOP218)Depreciation on fixed assets": ["Depreciation on fixed assets","Depreciation","DepreciationDepreciation and provisions","Depreciation and amortization"],
    "(AOP222)Other operating expenses": ["Other operating expenses","Other expenses","Intangible costs","Other expenses and losses"],
    "(AOP223)Income from financial transactions": ["Income from financial transactions (financial income)","III. FINANCIAL INCOME","Financial income"],
    "(AOP234)Financial costs": ["Financial costs"],
    "(AOP250+/AOP251-)Profit or loss before taxation": ["Profit or loss before taxation","Profit before taxation","Loss before taxation","Profit before taxation/Loss before taxation","Profit Before Tax"],
    "(AOP252)Profit tax": ["Profit tax","Income tax","Taxes,duties and similar expenses","Taxes","Tax","Tax charge","Tax expense of the period"],
    "(AOP255+/AOP256-)Profit or loss after taxation": ["Profit or loss after taxation","Profit after taxation","Loss after taxation","TOTAL RESULT"],
}

# Placeholders for other countries
allowed_fields_croatia = { 
    "(AOP001)Fixed assets": ["Fixed assets"],
    "(AOP002)Intangible assets": ["I. Intangible assets", "Intangible assets", "Intangible fixed assets", "Total Intangible assets"],
    "(AOP004)Concessions,patents,licenses and similar rights and other intangible assets":["Concessions,industrial rights,licences","Concessions, patents, licenses, trade marks etc.", "Service marks,software and other intangible assets","Concessions,patents,licenses and similar rights and other intangible assets","Concessions, patents, licenses, trademarks, service marks, software and other intangible assets"],
    "(AOP005)Goodwill":["Goodwill"],
    "(AOP006)Advances for intangible assets and intangible assets in preparation":["Advances for intangible assets and intangible assets in preparation"],
    "(AOP008)Other intangible fixed assets":["Other intangible fixed assets","Other Intangible Assets"],
    "(AOP009)Tangible fixed assets": ["II. Tangible assets","Tangible fixed assets", "Total tangible assets" , "Property,plants,equipment and biological assets", "Immovables,plants and equipment","Immovables, plants and equipment","Tangible assets","Property, plants, equipment and biological assets"],
    "(AOP010)Land and buildings":["Buildings"],
    "(AOP011)Land":["Land","Land and buildings"],
    "(AOP012)Buildings":["Buildings","Land and buildings"],
    "(AOP013)Machinery and equipment": ["Machinery and equipment", "Plant and Machinery", "Technical equipment and machinery", "Property,plant and equipment","Plant and equipment","Plant & Machinery"],
    "(AOP014)Other equipment, furniture, fittings, tools, fixtures, vehicles": [""],
    "(AOP016)Biological property":["Biological property","Biological assets"],
    "(AOP017)Advance payments for tangible assets": ["Advance payments for tangible assets", "Advances in property,plant,equipment and biological assets and property,biological asset in preparation","Advances in property, plant, equipment and biological assets and property, plant, equipment and biological assets in preparation"],
    "(AOP018)Tangible assets in progress": ["Tangible assets in progress", "Payments and fixed goods under construction","Immovables, plant  and equipment under construction"],
    "(AOP019)Other tangible assets": ["Other equipment, furniture, fittings, tools, fixtures, vehicles"],
    "(AOP020)Investments in real estate": ["Investments in real estate","Investment property measured at the cost model"],
    "(AOP021)Financial fixed assets": ["Financial fixed assets", "Long term investments","Long term financial investments and long term receivables Financial fixed assets","LONG-TERM FINANCIAL INVESTMENTS AND LONG-TERM RECEIVABLES","Financial assets"],
    "(AOP024)Loans to Group":["Long-term loans"],
    "(AOP025)Long term loans":["Long term loans","Long-term loans to legal entities with equity participation (excluding subsidiaries)"],
    "(AOP031)Long-term receivables": ["Long-term receivables", "Other long term investments and receivables","Long-term operating receivables"],
    "(AOP035)Deferred tax assets": ["DEFERRED TAX ASSETS", "Deferred tax assets"],
    "(AOP030)Other financial fixed assets":["Other financial fixed assets"],
    "(AOP034)Other long term receivables":["Other long term receivables","Other long-term receivables"],
    "(AOP036)Short-term assets": ["Short term assets", "Current assets", "Total Current assets", "Short term assets"],
    "(AOP037)Inventory": ["Inventory", "Total inventories", "Inventories"],
    "(AOP038)Raw materials, consumabeles and supplies":["Raw materials, consumabeles and supplies", "Inventory-raw materials, consumabeles and supplies","Inventory - raw materials, consumables, working inventory", "Inventory of materials(fabrication material, spare parts, small inventory and car tires)","Inventory of materials (fabrication material, spare parts, small inventory and car tires"],
    "(AOP039)Inventory of materials":["Inventory of materials(fabricationmaterial, spare parts)"],
    "(AOP040)Work in progress":["Work in progress"],
    "(AOP041)Finished products and goods":["Finished products and goods", "Finished products/merchandise", "Finished goods", "Goods"],
    "(AOP042)Traiding Goods":["Traiding Goods","Inventory-Trading Goods","Inventory - Trading Goods"],
    "(AOP048)Prepayments":["Prepayments","Payments in advance for stock"],
    "(AOP045)Short-term receivables": ["Short-term receivables"],
    "(AOP046)Short-term intercompany receivables": ["Short-term intercompany receivables", "Receivables from parent companies and subsidiaries", "Receivables from affiliated companies", "Intercompany Receivables","Receivables from foreign parent companies, subsidiaries and other associated companies","Group Receivables"],
    "(AOP047)Short-term trade receivables": ["Short-term trade receivables", "Receivables from buyers", "Trade debtor" , "Trade receivables","Trade debtors"],
    "(AOP050)Short-term receivables from employees": ["Short-term receivables from employees"],
    "(AOP049)Receivables from the state and other institutions": ["Prepaid corporate income tax","Receivables from the state and other institutions"],
    "(AOP051)Other short-term receivables": ["Other short-term receivables","Other short term receivables", "Other receivables", "Miscellaneous receivables"],
    "(AOP052)Short-term financial assets": ["SHORT TERM FINANCIAL ASSETS","short-term financial assets", "Financial investment", "Short term financial assets","Short-term financial investments","Short-term investments"],
    "(AOP058)Other short-term financial investments":["other short-term financial investments","Other short term financial assets"],
    "(AOP059_060)Cash": ["Cash","Cash and cash equivalents", "Cash assets", "Bank balance cheques and cash on hand", "Cash and cash equivalent", "Cash","Bank balance, cheques and cash on hand","Liquid assets"],
    "(AOP062)Prepaid expenses": ["Prepaid expenses","ACCRUALS"],
    "(AOP063)Total assets": ["TOTAL ASSETS", "Total assets", "Balance sheet total"],
    "(AOP064)Off balance sheet items": ["Off balance sheet items"],
    "(AOP065)Equity capital": ["Equity capital", "Owners equity"],
    "(AOP066)Subscribed and paid capital": ["Subscribed and paid capital", "Basic capital", "Shareholders equity","Called capital", "Share capital", "Called and share capital","Subscribed capital","Called up share capital (issued capital stock)"],
    "(AOP067)Emission premium":["Emission premium","Share premium"],
    "(AOP068)Own shares":["Own shares","Called up share capital","Called up share capital (issued capital stock)","Share capital","Subscribed capital unpaid"],
    "(AOP069)Subscribed unpaid capital":["Subscribed unpaid capital","Unpaid issued capital","Subscribed capital unpaid"],
    "(AOP071)Capital reserves": ["RESERVES FROM PROFIT"],
    "(AOP070)Revaluation reserves": ["Revaluation reserves","POSITIVE REVALUATION RESERVES AND UNREALIZED PROFIT FROM FINANCIAL ASSETS AND OTHER ELEMENTS OF OTHER COMPREHENSIVE INCOME"],
    "(AOP072)Legal reserves":["Legal reserves"],
    "(AOP074)Other reserves":["Other reserves"],
    "(AOP075+/AOP076-)Profit or loss carried forward": ["Profit or loss carried forward","Retained profit (earnings) of the year (net profits)","Accumulated retained earnings from previous periods","Retained profit", "Retained earnings brought forward", "Net retained profit/net acumalted loss", "Net profit or loss from previos years","Equity, net retained profits/net accumulated losses (balance sheet)","Net profit or loss from previous periods","Accumulated profit reserve","Retained earnings or loss"],
    "(AOP077+/AOP078-)Net profit or loss for the year": ["Net profit or loss for the year","Current period profit", "Loss from previos years", "Net profit for the period" , "Profit of the year","Net profit or loss for the year", "Net profit for the period","Retained earnings for the current year"],
    "(AOP081)Liabilities": ["Liabilities", "Total liabilities","Total debt","Total debts"],
    "(AOP082)Provision for risks and charges":["Provision for risks and charges","Provisions for risks and charges"],
    "(AOP083)Provisions for pensions and similar obligations":["Provisions for pensions and similar obligations"],
    "(AOP084)Other provisions":["Other provisions"],
    "(AOP085)Long-term liabilities": ["Long-term liabilities", "Long term liabilities", "Total Long term liabilities","Long-term financial and operating liabilities","Long-term provisions and long-term liabilities"],
    "(AOP086)Long-term liabilities to affiliates": ["Long-term liabilities to affiliates", "Group payables due after 1 year", "Long term liabilities to affiliates"],
    "(AOP087)Trade liabilities":["Trade liabilities"],
    "(AOP090)Long-term liabilities for loans": ["Long-term liabilities for loans"],
    "(AOP092)Other loans/finance due after 1 year":["Other loans/finance due after 1 year","Long term liabilities payable to financial institutions"],
    "(AOP093)Other long-term liabilities": ["Other long-term liabilities", "Other long term liabilities"],
    "(AOP095)Short-term liabilities": ["Short-term liabilities","IV. SHORT-TERM LIABILITIES","Short- term liabilities","Short term liabilities","SHORT- TERM LIABILITIES", "TOTAL CURRENT LIABILITIES", "Total current liabilities","Short- term liabilities and short term provisions"," Short- term financial and operating liabilities"],
    "(AOP096)Short-term liabilities to affiliates": ["Short-term liabilities to affiliates","Liabilities of disposal groups", " Trade payables-foreing parent company, subsidiares and other associated companies", " Liabilities to affiliated companies","Group payables"],
    "(AOP098)Prepayments, deposits and gurantee":["Prepayments, deposits and gurantee","Liabilities for securities","Prepayments, deposits and guarantees"],
    "(AOP103)Liabilities for loans, deposits, etc. to companies within the group": ["Liabilities for loans, deposits, etc. to companies within the group"],
    "(AOP097)Short-term trade creditors": ["Short-term trade creditors", "Liabilities to suppliers","Trade Payables","Trade creditors(acounts payable)", " Trade payables","Commitments towards suppliers","Trade payable"],
    "(AOP100)Short-term liabilities to employees": ["Short-term liabilities to employees","Liabilities towards employees"],
    "(AOP099)Short-term liabilities for taxes, contributions and other fees": ["Short-term liabilities for taxes, contributions and other fees"],
    "(AOP101)Liability for income tax":["Liability for income tax","S-term Liabilities for tax,contributions and other fees"],
    "(AOP104)Obligations for taken loans and credits":["Short-term liabilities for loans"],
    "(AOP108)Other short-term liabilities": ["Other short term liabilities", "Other operating and liabilities and other short term liabilities", "Other short-term liabilities ",  "Other  liabilities  including accruals",  "Other  liabilities","Other short-term liabilities","Miscellaneous Liabilities"],
    "(AOP109)Accruals and deferred income": ["Accruals and deferred income"],
    "(AOP111)Total liabilities and funds": ["Total liabilities and funds", "TOTAL EQUITY AND LIABILITIES "," TOTAL LIABILITIES","BALANCE SHEET TOTAL","TOTAL LIABILITIES","Balance sheet total"],
    "(AOP112)Off balance sheet items": ["Off balance sheet items"],
    "(AOP201)Turnover, sales revenue": ["Turnover, sales revenue"],
    "(AOP202)Revenues from sales": ["Income from sales (outside group)"],
    "(AOP203)Other income(other revenues)": ["Other operating income (outside the group)"],
    "(AOP206)Own work capitalized": ["Own work capitalized"],
    "(AOP207)Operating expenses": ["Operating expenses","Operating Costs"],
    "(AOP208)Material costs": ["Cost of raw materials and consumables"],
    "(AOP209)Cost of goods sold": ["Cost of goods sold","The purchase value of the goods sold","Cost of materials (type of expenditure format)","Cost of goods sold and the cost of materials"],
    "(AOP213)Staff costs": ["Staff costs (employee costs)"],
    "(AOP218)Depreciation on fixed assets": ["Depreciation on fixed assets","Depreciation","DepreciationDepreciation and provisions","Depreciation and amortization"],
    "(AOP222)Other operating expenses": ["Other costs"],
    "(AOP223)Income from financial transactions": ["Income from financial transactions (financial income)","III. FINANCIAL INCOME","Financial income"],
    "(AOP234)Financial costs": ["Financial costs"],
    "(AOP250+/AOP251-)Profit or loss before taxation": ["Profit or loss before taxation","Profit before taxation","Loss before taxation","Profit before taxation/Loss before taxation","Profit Before Tax"],
    "(AOP252)Profit tax": ["Profit tax","Income tax","Taxes,duties and similar expenses","Taxes","Tax","Tax charge","Tax expense of the period"],
    "(AOP255+/AOP256-)Profit or loss after taxation": ["Profit or Loss after taxation"],
}
allowed_fields_albania = { 
    "(AOP001)Fixed assets": ["Fixed assets"],
    "(AOP002)Intangible assets": ["I. Intangible assets", "Intangible assets", "Intangible fixed assets", "Total Intangible assets"],
    "(AOP004)Concessions,patents,licenses and similar rights and other intangible assets":["Concessions,industrial rights,licences","Concessions, patents, licenses, trade marks etc.", "Service marks,software and other intangible assets","Concessions,patents,licenses and similar rights and other intangible assets","Concessions, patents, licenses, trademarks, service marks, software and other intangible assets"],
    "(AOP005)Goodwill":["Goodwill"],
    "(AOP006)Advances for intangible assets and intangible assets in preparation":["Advances for intangible assets and intangible assets in preparation"],
    "(AOP008)Other intangible fixed assets":["Other intangible fixed assets","Other Intangible Assets"],
    "(AOP009)Tangible fixed assets": ["II. Tangible assets","Tangible fixed assets", "Total tangible assets" , "Property,plants,equipment and biological assets", "Immovables,plants and equipment","Immovables, plants and equipment","Tangible assets","Property, plants, equipment and biological assets"],
    "(AOP010)Land and buildings":["Land and buildings","Property, plant and equipment","Land & Buildings"],
    "(AOP011)Land":["Land","Land and buildings"],
    "(AOP012)Buildings":["Buildings","Land and buildings"],
    "(AOP013)Machinery and equipment": ["Machinery and equipment", "Plant and Machinery", "Technical equipment and machinery", "Property,plant and equipment","Plant and equipment","Plant & Machinery"],
    "(AOP014)Other equipment, furniture, fittings, tools, fixtures, vehicles": ["Other equipment, furniture, fittings, tools, fixtures, vehicles"],
    "(AOP016)Biological property":["Biological property","Biological assets"],
    "(AOP017)Advance payments for tangible assets": ["Prepaid expenses"],
    "(AOP018)Tangible assets in progress": ["Tangible assets in progress", "Payments and fixed goods under construction","Immovables, plant  and equipment under construction"],
    "(AOP019)Other tangible assets": [""],
    "(AOP020)Investments in real estate": ["Investments in real estate","Investment property measured at the cost model"],
    "(AOP021)Financial fixed assets": ["Financial fixed assets", "Long term investments","Long term financial investments and long term receivables Financial fixed assets","LONG-TERM FINANCIAL INVESTMENTS AND LONG-TERM RECEIVABLES","Financial assets"],
    "(AOP024)Loans to Group":["Loans to Group","Long-term loans to parent and subsidiary legal entities"],
    "(AOP025)Long term loans":["Long term loans","Long-term loans to legal entities with equity participation (excluding subsidiaries)"],
    "(AOP031)Long-term receivables": ["Long-term receivables", "Other long term investments and receivables","Long-term operating receivables"],
    "(AOP035)Deferred tax assets": ["DEFERRED TAX ASSETS", "Deferred tax assets"],
    "(AOP030)Other financial fixed assets":["Other financial fixed assets"],
    "(AOP034)Other long term receivables":["Other long term receivables","Other long-term receivables"],
    "(AOP036)Short-term assets": ["Short term assets", "Current assets", "Total Current assets", "Short term assets"],
    "(AOP037)Inventory": ["Inventory", "Total inventories", "Inventories"],
    "(AOP038)Raw materials, consumabeles and supplies":["Raw materials, consumabeles and supplies", "Inventory-raw materials"],
    "(AOP039)Inventory of materials":["Inventory of materials(fabricationmaterial, spare parts)"],
    "(AOP040)Work in progress":["Work in progress"],
    "(AOP041)Finished products and goods":["Inventory - Finished goods"],
    "(AOP042)Traiding Goods":["Traiding Goods","Inventory-Trading Goods","Inventory - Trading Goods"],
    "(AOP048)Prepayments":["Prepayments","Payments in advance for stock"],
    "(AOP045)Short-term receivables": [""],
    "(AOP046)Short-term intercompany receivables": ["Short-term intercompany receivables", "Receivables from parent companies and subsidiaries", "Receivables from affiliated companies", "Intercompany Receivables","Receivables from foreign parent companies, subsidiaries and other associated companies","Group Receivables"],
    "(AOP047)Short-term trade receivables": ["Trade debtors (receivables from customer)"],
    "(AOP050)Short-term receivables from employees": ["Short-term receivables from employees"],
    "(AOP049)Receivables from the state and other institutions": ["Prepaid corporate income tax","Receivables from the state and other institutions"],
    "(AOP051)Other short-term receivables": ["Other short-term receivables","Other short term receivables", "Other receivables", "Miscellaneous receivables"],
    "(AOP052)Short-term financial assets": ["SHORT TERM FINANCIAL ASSETS","short-term financial assets", "Financial investment", "Short term financial assets","Short-term financial investments","Short-term investments"],
    "(AOP058)Other short-term financial investments":["other short-term financial investments","Other short term financial assets"],
    "(AOP059_060)Cash": ["Cash","Cash and cash equivalents", "Cash assets", "Bank balance cheques and cash on hand", "Cash and cash equivalent", "Cash","Bank balance, cheques and cash on hand","Liquid assets"],
    "(AOP062)Prepaid expenses": [""],
    "(AOP063)Total assets": ["TOTAL ASSETS", "Total assets", "Balance sheet total"],
    "(AOP064)Off balance sheet items": ["Off balance sheet items"],
    "(AOP065)Equity capital": ["Equity capital", "Owners equity"],
    "(AOP066)Subscribed and paid capital": ["Subscribed and paid capital", "Basic capital", "Shareholders equity","Called capital", "Share capital", "Called and share capital","Subscribed capital","Called up share capital (issued capital stock)"],
    "(AOP067)Emission premium":["Emission premium","Share premium"],
    "(AOP068)Own shares":["Own shares","Called up share capital","Called up share capital (issued capital stock)","Share capital","Subscribed capital unpaid"],
    "(AOP069)Subscribed unpaid capital":["Subscribed unpaid capital","Unpaid issued capital","Subscribed capital unpaid"],
    "(AOP071)Capital reserves": ["RESERVES FROM PROFIT"],
    "(AOP070)Revaluation reserves": ["Revaluation reserves","POSITIVE REVALUATION RESERVES AND UNREALIZED PROFIT FROM FINANCIAL ASSETS AND OTHER ELEMENTS OF OTHER COMPREHENSIVE INCOME"],
    "(AOP072)Legal reserves":["Legal reserves"],
    "(AOP074)Other reserves":["Other reserves"],
    "(AOP075+/AOP076-)Profit or loss carried forward": ["Profit or loss carried forward","Retained profit (earnings) of the year (net profits)","Accumulated retained earnings from previous periods","Retained profit", "Retained earnings brought forward", "Net retained profit/net acumalted loss", "Net profit or loss from previos years","Equity, net retained profits/net accumulated losses (balance sheet)","Net profit or loss from previous periods","Accumulated profit reserve","Retained earnings or loss"],
    "(AOP077+/AOP078-)Net profit or loss for the year": ["Net profit or loss for the year","Current period profit", "Loss from previos years", "Net profit for the period" , "Profit of the year","Net profit or loss for the year", "Net profit for the period","Retained earnings for the current year"],
    "(AOP081)Liabilities": ["TOTAL LIABILITIES"],
    "(AOP082)Provision for risks and charges":["Provision for risks and charges","Provisions for risks and charges"],
    "(AOP083)Provisions for pensions and similar obligations":["Provisions for pensions and similar obligations"],
    "(AOP084)Other provisions":["Other provisions"],
    "(AOP085)Long-term liabilities": ["Long-term liabilities", "Long term liabilities", "Total Long term liabilities","Long-term financial and operating liabilities","Long-term provisions and long-term liabilities"],
    "(AOP086)Long-term liabilities to affiliates": ["Long-term liabilities to affiliates", "Group payables due after 1 year", "Long term liabilities to affiliates"],
    "(AOP087)Trade liabilities":["Trade liabilities"],
    "(AOP090)Long-term liabilities for loans": ["Long-term liabilities for loans"],
    "(AOP092)Other loans/finance due after 1 year":["Other loans/finance due after 1 year","Long term liabilities payable to financial institutions"],
    "(AOP093)Other long-term liabilities": ["Other long-term liabilities", "Other long term liabilities"],
    "(AOP095)Short-term liabilities": ["Short-term liabilities","IV. SHORT-TERM LIABILITIES","Short- term liabilities","Short term liabilities","SHORT- TERM LIABILITIES", "TOTAL CURRENT LIABILITIES", "Total current liabilities","Short- term liabilities and short term provisions"," Short- term financial and operating liabilities"],
    "(AOP096)Short-term liabilities to affiliates": ["Short-term liabilities to affiliates","Liabilities of disposal groups", " Trade payables-foreing parent company, subsidiares and other associated companies", " Liabilities to affiliated companies","Group payables"],
    "(AOP098)Prepayments, deposits and gurantee":["Prepayments, deposits and gurantee","Liabilities for securities","Prepayments, deposits and guarantees"],
    "(AOP103)Liabilities for loans, deposits, etc. to companies within the group": ["Liabilities for loans, deposits, etc. to companies within the group"],
    "(AOP097)Short-term trade creditors": ["Short-term trade creditors", "Liabilities to suppliers","Trade Payables","Trade creditors(acounts payable)", " Trade payables","Commitments towards suppliers","Trade payable"],
    "(AOP100)Short-term liabilities to employees": ["Short-term liabilities to employees","Liabilities towards employees"],
    "(AOP099)Short-term liabilities for taxes, contributions and other fees": ["Short-term liabilities for taxes, contributions and other fees"],
    "(AOP101)Liability for income tax":["Liability for income tax","S-term Liabilities for tax,contributions and other fees"],
    "(AOP104)Obligations for taken loans and credits":["Short-term loans"],
    "(AOP108)Other short-term liabilities": ["Other short term liabilities", "Other operating and liabilities and other short term liabilities", "Other short-term liabilities ",  "Other  liabilities  including accruals",  "Other  liabilities","Other short-term liabilities","Miscellaneous Liabilities"],
    "(AOP109)Accruals and deferred income": ["Accruals and deferred income"],
    "(AOP111)Total liabilities and funds": ["Total liabilities and funds"],
    "(AOP112)Off balance sheet items": ["Off balance sheet items"],
    "(AOP201)Turnover, sales revenue": ["Turnover, sales revenue"],
    "(AOP202)Revenues from sales": ["Revenues from sales"],
    "(AOP203)Other income(other revenues)": ["Other income (other revenues)"],
    "(AOP206)Own work capitalized": ["Own work capitalized"],
    "(AOP207)Operating expenses": ["Operating expenses","Operating Costs"],
    "(AOP208)Material costs": ["Cost of raw materials and consumables"],
    "(AOP209)Cost of goods sold": ["Cost of goods sold","The purchase value of the goods sold","Cost of materials (type of expenditure format)","Cost of goods sold and the cost of materials"],
    "(AOP213)Staff costs": ["Staff costs (employee costs)"],
    "(AOP218)Depreciation on fixed assets": ["Depreciation on fixed assets","Depreciation","DepreciationDepreciation and provisions","Depreciation and amortization"],
    "(AOP222)Other operating expenses": ["Other costs"],
    "(AOP223)Income from financial transactions": ["Income from financial transactions (financial income)","III. FINANCIAL INCOME","Financial income"],
    "(AOP234)Financial costs": ["Financial costs"],
    "(AOP250+/AOP251-)Profit or loss before taxation": ["Profit or loss before taxation","Profit before taxation","Loss before taxation","Profit before taxation/Loss before taxation","Profit Before Tax"],
    "(AOP252)Profit tax": ["Profit tax","Income tax","Taxes,duties and similar expenses","Taxes","Tax","Tax charge","Tax expense of the period"],
    "(AOP255+/AOP256-)Profit or loss after taxation": ["Profit after taxation"],
}
allowed_fields_belgium = { 
    "(AOP001)Fixed assets": ["Fixed assets"],
    "(AOP002)Intangible assets": ["I. Intangible assets", "Intangible assets", "Intangible fixed assets", "Total Intangible assets"],
    "(AOP004)Concessions,patents,licenses and similar rights and other intangible assets":["Concessions,industrial rights,licences","Concessions, patents, licenses, trade marks etc.", "Service marks,software and other intangible assets","Concessions,patents,licenses and similar rights and other intangible assets","Concessions, patents, licenses, trademarks, service marks, software and other intangible assets"],
    "(AOP005)Goodwill":["Goodwill"],
    "(AOP006)Advances for intangible assets and intangible assets in preparation":["Advances for intangible assets and intangible assets in preparation"],
    "(AOP008)Other intangible fixed assets":["Other intangible fixed assets","Other Intangible Assets"],
    "(AOP009)Tangible fixed assets": ["II. Tangible assets","Tangible fixed assets", "Total tangible assets" , "Property,plants,equipment and biological assets", "Immovables,plants and equipment","Immovables, plants and equipment","Tangible assets","Property, plants, equipment and biological assets"],
    "(AOP010)Land and buildings":["Buildings"],
    "(AOP011)Land":["Land","Land and buildings"],
    "(AOP012)Buildings":["Buildings","Land and buildings"],
    "(AOP013)Machinery and equipment": ["Machinery and equipment", "Plant and Machinery", "Technical equipment and machinery", "Property,plant and equipment","Plant and equipment","Plant & Machinery"],
    "(AOP014)Other equipment, furniture, fittings, tools, fixtures, vehicles": [""],
    "(AOP016)Biological property":["Biological property","Biological assets"],
    "(AOP017)Advance payments for tangible assets": ["Advance payments for tangible assets", "Advances in property,plant,equipment and biological assets and property,biological asset in preparation","Advances in property, plant, equipment and biological assets and property, plant, equipment and biological assets in preparation"],
    "(AOP018)Tangible assets in progress": ["Tangible assets in progress", "Payments and fixed goods under construction","Immovables, plant  and equipment under construction"],
    "(AOP019)Other tangible assets": ["Other Tangible Assets"],
    "(AOP020)Investments in real estate": ["Investments in real estate","Investment property measured at the cost model"],
    "(AOP021)Financial fixed assets": ["Financial fixed assets", "Long term investments","Long term financial investments and long term receivables Financial fixed assets","LONG-TERM FINANCIAL INVESTMENTS AND LONG-TERM RECEIVABLES","Financial assets"],
    "(AOP024)Loans to Group":["Long-term loans"],
    "(AOP025)Long term loans":["Long term loans","Long-term loans to legal entities with equity participation (excluding subsidiaries)"],
    "(AOP031)Long-term receivables": ["Long-term receivables", "Other long term investments and receivables","Long-term operating receivables"],
    "(AOP035)Deferred tax assets": ["DEFERRED TAX ASSETS", "Deferred tax assets"],
    "(AOP030)Other financial fixed assets":["Other financial fixed assets"],
    "(AOP034)Other long term receivables":["Other long term receivables","Other long-term receivables"],
    "(AOP036)Short-term assets": ["Short term assets", "Current assets", "Total Current assets", "Short term assets"],
    "(AOP037)Inventory": ["Inventory", "Total inventories", "Inventories"],
    "(AOP038)Raw materials, consumabeles and supplies":["Raw materials, consumabeles and supplies", "Inventory-raw materials, consumabeles and supplies","Inventory - raw materials, consumables, working inventory", "Inventory of materials(fabrication material, spare parts, small inventory and car tires)","Inventory of materials (fabrication material, spare parts, small inventory and car tires"],
    "(AOP039)Inventory of materials":["Inventory of materials(fabricationmaterial, spare parts)"],
    "(AOP040)Work in progress":["Work in progress"],
    "(AOP041)Finished products and goods":["Finished products and goods", "Finished products/merchandise", "Finished goods", "Goods"],
    "(AOP042)Traiding Goods":["Traiding Goods","Inventory-Trading Goods","Inventory - Trading Goods"],
    "(AOP048)Prepayments":["Prepayments","Payments in advance for stock"],
    "(AOP045)Short-term receivables": ["Total Receivables"],
    "(AOP046)Short-term intercompany receivables": ["Short-term intercompany receivables", "Receivables from parent companies and subsidiaries", "Receivables from affiliated companies", "Intercompany Receivables","Receivables from foreign parent companies, subsidiaries and other associated companies","Group Receivables"],
    "(AOP047)Short-term trade receivables": ["Short-term trade receivables", "Receivables from buyers", "Trade debtor" , "Trade receivables","Trade debtors"],
    "(AOP050)Short-term receivables from employees": ["Short-term receivables from employees"],
    "(AOP049)Receivables from the state and other institutions": ["Prepaid corporate income tax","Receivables from the state and other institutions"],
    "(AOP051)Other short-term receivables": ["Other short-term receivables","Other short term receivables", "Other receivables", "Miscellaneous receivables"],
    "(AOP052)Short-term financial assets": ["SHORT TERM FINANCIAL ASSETS","short-term financial assets", "Financial investment", "Short term financial assets","Short-term financial investments","Short-term investments"],
    "(AOP058)Other short-term financial investments":["Other Current Assets"],
    "(AOP059_060)Cash": ["Cash","Cash and cash equivalents", "Cash assets", "Bank balance cheques and cash on hand", "Cash and cash equivalent", "Cash","Bank balance, cheques and cash on hand","Liquid assets"],
    "(AOP062)Prepaid expenses": ["Prepaid expenses","ACCRUALS"],
    "(AOP063)Total assets": [""],
    "(AOP064)Off balance sheet items": ["Balance sheet total"],
    "(AOP065)Equity capital": ["Equity capital", "Owners equity"],
    "(AOP066)Subscribed and paid capital": ["Subscribed and paid capital", "Basic capital", "Shareholders equity","Called capital", "Share capital", "Called and share capital","Subscribed capital","Called up share capital (issued capital stock)"],
    "(AOP067)Emission premium":["Emission premium","Share premium"],
    "(AOP068)Own shares":[""],
    "(AOP069)Subscribed unpaid capital":["Subscribed unpaid capital","Unpaid issued capital","Subscribed capital unpaid"],
    "(AOP071)Capital reserves": ["Revenue Reserves"],
    "(AOP070)Revaluation reserves": ["Revaluation reserves","POSITIVE REVALUATION RESERVES AND UNREALIZED PROFIT FROM FINANCIAL ASSETS AND OTHER ELEMENTS OF OTHER COMPREHENSIVE INCOME"],
    "(AOP072)Legal reserves":["Legal reserves"],
    "(AOP074)Other reserves":["Other reserves"],
    "(AOP075+/AOP076-)Profit or loss carried forward": ["Profit or loss carried forward","Retained profit (earnings) of the year (net profits)","Accumulated retained earnings from previous periods","Retained profit", "Retained earnings brought forward", "Net retained profit/net acumalted loss", "Net profit or loss from previos years","Equity, net retained profits/net accumulated losses (balance sheet)","Net profit or loss from previous periods","Accumulated profit reserve","Retained earnings or loss"],
    "(AOP077+/AOP078-)Net profit or loss for the year": ["Net profit or loss for the year","Current period profit", "Loss from previos years", "Net profit for the period" , "Profit of the year","Net profit or loss for the year", "Net profit for the period","Retained earnings for the current year"],
    "(AOP081)Liabilities": ["Liabilities", "Total liabilities","Total debt","Total debts"],
    "(AOP082)Provision for risks and charges":["Provision for risks and charges","Provisions for risks and charges"],
    "(AOP083)Provisions for pensions and similar obligations":["Provisions for pensions and similar obligations"],
    "(AOP084)Other provisions":["Other provisions"],
    "(AOP085)Long-term liabilities": ["Long-term liabilities", "Long term liabilities", "Total Long term liabilities","Long-term financial and operating liabilities","Long-term provisions and long-term liabilities"],
    "(AOP086)Long-term liabilities to affiliates": ["Long-term liabilities to affiliates", "Group payables due after 1 year", "Long term liabilities to affiliates"],
    "(AOP087)Trade liabilities":["Trade liabilities"],
    "(AOP090)Long-term liabilities for loans": ["Other loans/finance due after 1 year"],
    "(AOP092)Other loans/finance due after 1 year":[""],
    "(AOP093)Other long-term liabilities": ["Other long-term liabilities", "Other long term liabilities"],
    "(AOP095)Short-term liabilities": ["Short-term liabilities","IV. SHORT-TERM LIABILITIES","Short- term liabilities","Short term liabilities","SHORT- TERM LIABILITIES", "TOTAL CURRENT LIABILITIES", "Total current liabilities","Short- term liabilities and short term provisions"," Short- term financial and operating liabilities"],
    "(AOP096)Short-term liabilities to affiliates": ["Short-term liabilities to affiliates","Liabilities of disposal groups", " Trade payables-foreing parent company, subsidiares and other associated companies", " Liabilities to affiliated companies","Group payables"],
    "(AOP098)Prepayments, deposits and gurantee":["Prepayments, deposits and gurantee","Liabilities for securities","Prepayments, deposits and guarantees"],
    "(AOP103)Liabilities for loans, deposits, etc. to companies within the group": ["Liabilities for loans, deposits, etc. to companies within the group"],
    "(AOP097)Short-term trade creditors": ["Short-term trade creditors", "Liabilities to suppliers","Trade Payables","Trade creditors(acounts payable)", " Trade payables","Commitments towards suppliers","Trade payable"],
    "(AOP100)Short-term liabilities to employees": ["Short-term liabilities to employees","Liabilities towards employees"],
    "(AOP099)Short-term liabilities for taxes, contributions and other fees": ["Short-term liabilities for taxes, contributions and other fees"],
    "(AOP101)Liability for income tax":["Liability for income tax","S-term Liabilities for tax,contributions and other fees"],
    "(AOP104)Obligations for taken loans and credits":["Other Loans/Finance"],
    "(AOP108)Other short-term liabilities": ["Other short term liabilities", "Other operating and liabilities and other short term liabilities", "Other short-term liabilities ",  "Other  liabilities  including accruals",  "Other  liabilities","Other short-term liabilities","Miscellaneous Liabilities"],
    "(AOP109)Accruals and deferred income": ["Accruals and deferred income"],
    "(AOP111)Total liabilities and funds": [""],
    "(AOP112)Off balance sheet items": ["Off balance sheet items"],
    "(AOP201)Turnover, sales revenue": ["Turnover"],
    "(AOP202)Revenues from sales": ["Income from sales (outside group)"],
    "(AOP203)Other income(other revenues)": ["Other operating income (outside the group)"],
    "(AOP206)Own work capitalized": ["Own work capitalized"],
    "(AOP207)Operating expenses": ["Operating expenses","Operating Costs"],
    "(AOP208)Material costs": ["Cost of raw materials and consumables"],
    "(AOP209)Cost of goods sold": ["Cost of goods sold","The purchase value of the goods sold","Cost of materials (type of expenditure format)","Cost of goods sold and the cost of materials"],
    "(AOP213)Staff costs": ["Wages & Salaries"],
    "(AOP218)Depreciation on fixed assets": ["Depreciation on fixed assets","Depreciation","DepreciationDepreciation and provisions","Depreciation and amortization"],
    "(AOP222)Other operating expenses": ["Other costs"],
    "(AOP223)Income from financial transactions": ["Income from financial transactions (financial income)","III. FINANCIAL INCOME","Financial income"],
    "(AOP234)Financial costs": ["Financial Expenses"],
    "(AOP250+/AOP251-)Profit or loss before taxation": ["Profit Before Tax"],
    "(AOP252)Profit tax": ["Tax"],
    "(AOP255+/AOP256-)Profit or loss after taxation": ["Profit after tax"],
}
allowed_fields_bulgaria = {
    "(AOP001)Fixed assets": ["Fixed assets"],
    "(AOP002)Intangible assets": ["I. Intangible assets", "Intangible assets", "Intangible fixed assets", "Total Intangible assets"],
    "(AOP004)Concessions,patents,licenses and similar rights and other intangible assets":["Concessions,industrial rights,licences","Concessions, patents, licenses, trade marks etc.", "Service marks,software and other intangible assets","Concessions,patents,licenses and similar rights and other intangible assets","Concessions, patents, licenses, trademarks, service marks, software and other intangible assets"],
    "(AOP005)Goodwill":["Goodwill"],
    "(AOP006)Advances for intangible assets and intangible assets in preparation":["Advances for intangible assets and intangible assets in preparation"],
    "(AOP008)Other intangible fixed assets":["Other intangible fixed assets","Other Intangible Assets"],
    "(AOP009)Tangible fixed assets": ["II. Tangible assets","Tangible fixed assets", "Total tangible assets" , "Property,plants,equipment and biological assets", "Immovables,plants and equipment","Immovables, plants and equipment","Tangible assets","Property, plants, equipment and biological assets"],
    "(AOP010)Land and buildings":["Land and buildings","Property, plant and equipment","Land & Buildings"],
    "(AOP011)Land":["Land","Land and buildings"],
    "(AOP012)Buildings":["Buildings","Land and buildings"],
    "(AOP013)Machinery and equipment": ["Machinery and equipment", "Plant and Machinery", "Technical equipment and machinery", "Property,plant and equipment","Plant and equipment","Plant & Machinery"],
    "(AOP014)Other equipment, furniture, fittings, tools, fixtures, vehicles": ["Other equipment, furniture, fittings, tools, fixtures, vehicles"],
    "(AOP016)Biological property":["Biological assets"],
    "(AOP017)Advance payments for tangible assets": ["Advance payments for tangible assets", "Advances in property,plant,equipment and biological assets and property,biological asset in preparation","Advances in property, plant, equipment and biological assets and property, plant, equipment and biological assets in preparation"],
    "(AOP018)Tangible assets in progress": ["Tangible assets in progress", "Payments and fixed goods under construction","Immovables, plant  and equipment under construction"],
    "(AOP019)Other tangible assets": ["Other tangible assets","Other tangible fixed assets", "Other equipment" ,"Other Immovables,plants and equipment","Other immovables, plant and equipment and investment in third-party immovables, plant and equipment","Other equipment, furniture, fittings, tools, fixtures, vehicles","Other unspecified material fixed assets"],
    "(AOP020)Investments in real estate": ["Investment property"],
    "(AOP021)Financial fixed assets": ["Financial fixed assets", "Long term investments","Long term financial investments and long term receivables Financial fixed assets","LONG-TERM FINANCIAL INVESTMENTS AND LONG-TERM RECEIVABLES","Financial assets"],
    "(AOP024)Loans to Group":["Loans to Group","Long-term loans to parent and subsidiary legal entities"],
    "(AOP025)Long term loans":["Long term loans","Long-term loans to legal entities with equity participation (excluding subsidiaries)"],
    "(AOP031)Long-term receivables": ["Long-term receivables", "Other long term investments and receivables","Long-term operating receivables"],
    "(AOP035)Deferred tax assets": ["DEFERRED TAX ASSETS", "Deferred tax assets"],
    "(AOP030)Other financial fixed assets":["Other financial fixed assets"],
    "(AOP034)Other long term receivables":["Other long term receivables","Other long-term receivables"],
    "(AOP036)Short-term assets": ["Short term assets", "Current assets", "Total Current assets", "Short term assets"],
    "(AOP037)Inventory": ["Inventory", "Total inventories", "Inventories"],
    "(AOP038)Raw materials, consumabeles and supplies":["Raw materials, consumabeles and supplies", "Inventory-raw materials, consumabeles and supplies","Inventory - raw materials, consumables, working inventory", "Inventory of materials(fabrication material, spare parts, small inventory and car tires)","Inventory of materials (fabrication material, spare parts, small inventory and car tires"],
    "(AOP039)Inventory of materials":["Inventory of materials(fabricationmaterial, spare parts)"],
    "(AOP040)Work in progress":["Work in progress"],
    "(AOP041)Finished products and goods":["Finished products and goods", "Finished products/merchandise", "Finished goods", "Goods"],
    "(AOP042)Traiding Goods":["Traiding Goods","Inventory-Trading Goods","Inventory - Trading Goods"],
    "(AOP048)Prepayments":["Prepayments","Payments in advance for stock"],
    "(AOP045)Short-term receivables": ["Short-term receivables"],
    "(AOP046)Short-term intercompany receivables": ["Short-term intercompany receivables", "Receivables from parent companies and subsidiaries", "Receivables from affiliated companies", "Intercompany Receivables","Receivables from foreign parent companies, subsidiaries and other associated companies","Group Receivables"],
    "(AOP047)Short-term trade receivables": ["Trade debtors (receivables from customer)"],
    "(AOP050)Short-term receivables from employees": ["Short-term receivables from employees"],
    "(AOP049)Receivables from the state and other institutions": ["Prepaid corporate income tax","Receivables from the state and other institutions"],
    "(AOP051)Other short-term receivables": ["Other debtors (other receivables)"],
    "(AOP052)Short-term financial assets": ["SHORT TERM FINANCIAL ASSETS","short-term financial assets", "Financial investment", "Short term financial assets","Short-term financial investments","Short-term investments"],
    "(AOP058)Other short-term financial investments":["other short-term financial investments","Other short term financial assets"],
    "(AOP059_060)Cash": ["Cash","Cash and cash equivalents", "Cash assets", "Bank balance cheques and cash on hand", "Cash and cash equivalent", "Cash","Bank balance, cheques and cash on hand","Liquid assets"],
    "(AOP062)Prepaid expenses": ["Prepaid expenses, deferred income, similar accounts"],
    "(AOP063)Total assets": ["TOTAL ASSETS", "Total assets", "Balance sheet total"],
    "(AOP064)Off balance sheet items": ["Off balance sheet items"],
    "(AOP065)Equity capital": ["Equity capital", "Owners equity"],
    "(AOP066)Subscribed and paid capital": ["Shareholders equity"],
    "(AOP067)Emission premium":["Subscribed and paid capital"],
    "(AOP068)Own shares":["Own shares","Called up share capital","Called up share capital (issued capital stock)","Share capital","Subscribed capital unpaid"],
    "(AOP069)Subscribed unpaid capital":["Subscribed unpaid capital","Unpaid issued capital","Subscribed capital unpaid"],
    "(AOP071)Capital reserves": ["CAPITAL RESERVES", "Capital reserves", "Reserves", "Revenue","Revenue Reserves","Capital reserve"],
    "(AOP070)Revaluation reserves": ["Revaluation reserves","POSITIVE REVALUATION RESERVES AND UNREALIZED PROFIT FROM FINANCIAL ASSETS AND OTHER ELEMENTS OF OTHER COMPREHENSIVE INCOME"],
    "(AOP072)Legal reserves":["Legal reserves"],
    "(AOP074)Other reserves":["Other reserves"],
    "(AOP075+/AOP076-)Profit or loss carried forward": ["Profit or loss carried forward","Retained profit (earnings) of the year (net profits)","Accumulated retained earnings from previous periods","Retained profit", "Retained earnings brought forward", "Net retained profit/net acumalted loss", "Net profit or loss from previos years","Equity, net retained profits/net accumulated losses (balance sheet)","Net profit or loss from previous periods","Accumulated profit reserve","Retained earnings or loss"],
    "(AOP077+/AOP078-)Net profit or loss for the year": ["Net profit or loss for the year","Current period profit", "Loss from previos years", "Net profit for the period" , "Profit of the year","Net profit or loss for the year", "Net profit for the period","Retained earnings for the current year"],
    "(AOP081)Liabilities": ["Liabilities", "Total liabilities","Total debt","Total debts"],
    "(AOP082)Provision for risks and charges":["Provision for risks and charges","Provisions for risks and charges"],
    "(AOP083)Provisions for pensions and similar obligations":["Provisions for pensions and similar obligations"],
    "(AOP084)Other provisions":["Other provisions"],
    "(AOP085)Long-term liabilities": ["Long-term liabilities", "Long term liabilities", "Total Long term liabilities","Long-term financial and operating liabilities","Long-term provisions and long-term liabilities"],
    "(AOP086)Long-term liabilities to affiliates": ["Long-term liabilities to affiliates", "Group payables due after 1 year", "Long term liabilities to affiliates"],
    "(AOP087)Trade liabilities":["Trade liabilities"],
    "(AOP090)Long-term liabilities for loans": ["L-term Liabilities payable to financial institutions"],
    "(AOP092)Other loans/finance due after 1 year":["Other loans/finance due after 1 year","Long term liabilities payable to financial institutions"],
    "(AOP093)Other long-term liabilities": ["Other long-term liabilities", "Other long term liabilities"],
    "(AOP095)Short-term liabilities": ["Short-term liabilities","IV. SHORT-TERM LIABILITIES","Short- term liabilities","Short term liabilities","SHORT- TERM LIABILITIES", "TOTAL CURRENT LIABILITIES", "Total current liabilities","Short- term liabilities and short term provisions"," Short- term financial and operating liabilities"],
    "(AOP096)Short-term liabilities to affiliates": ["Short-term liabilities to affiliates","Liabilities of disposal groups", " Trade payables-foreing parent company, subsidiares and other associated companies", " Liabilities to affiliated companies","Group payables"],
    "(AOP098)Prepayments, deposits and gurantee":["Prepayments, deposits and gurantee","Liabilities for securities","Prepayments, deposits and guarantees"],
    "(AOP103)Liabilities for loans, deposits, etc. to companies within the group": ["Liabilities for loans, deposits, etc. to companies within the group"],
    "(AOP097)Short-term trade creditors": ["Trade creditors (accounts payable)"],
    "(AOP100)Short-term liabilities to employees": ["Short-term liabilities to employees","Liabilities towards employees"],
    "(AOP099)Short-term liabilities for taxes, contributions and other fees": ["Short-term liabilities for taxes, contributions and other fees"],
    "(AOP101)Liability for income tax":["S-term Liabilities for taxes, contributions and other fees"],
    "(AOP104)Obligations for taken loans and credits":["Obligations for taken loans and credits","Short Term Bank Debt","Short liabilities for loans","Short term loans","Short term financial liabilities","Obligation for taken loans","BANK LOANS AND OVERDRAFT","Liabilities to bank","Loans liabilities from credit institutions","Short-term liabilities for loans","Short-term financial liabilities","S-term Liabilities payable to financial institutions","Banks loans and overdraft","Other Loans/Finance"],
    "(AOP108)Other short-term liabilities": ["Other short term liabilities", "Other operating and liabilities and other short term liabilities", "Other short-term liabilities ",  "Other  liabilities  including accruals",  "Other  liabilities","Other short-term liabilities","Miscellaneous Liabilities"],
    "(AOP109)Accruals and deferred income": ["Accruals and deferred income"],
    "(AOP111)Total liabilities and funds": ["Total liabilities and funds", "TOTAL EQUITY AND LIABILITIES "," TOTAL LIABILITIES","BALANCE SHEET TOTAL","TOTAL LIABILITIES","Balance sheet total"],
    "(AOP112)Off balance sheet items": ["Off balance sheet items"],
    "(AOP201)Turnover, sales revenue": ["Turnover"],
    "(AOP202)Revenues from sales": [""],
    "(AOP203)Other income(other revenues)": ["Other income(other revenues)","Other operating income (outside the group)","Other operating income","OTHER OPERATING REVENUE","Other income and profits","Income from financial transactions (financial income)","Other income"],
    "(AOP206)Own work capitalized": ["Own work capitalized"],
    "(AOP207)Operating expenses": ["Operating expenses","Operating Costs"],
    "(AOP208)Material costs": ["Material costs","Cost of raw materials and consumables","RAW MATERIAL COSTS, FUEL AND ENERGY COSTS"],
    "(AOP209)Cost of goods sold": ["Cost of goods sold","The purchase value of the goods sold","Cost of materials (type of expenditure format)","Cost of goods sold and the cost of materials"],
    "(AOP213)Staff costs": ["Staff costs (employee costs)","Wages and salaries","Wages & Salaries","Short-term financial and operating liabilities","Personnel type expenses","Wages expenses, wage compensation and other personal expenses"],
    "(AOP218)Depreciation on fixed assets": ["Depreciation on fixed assets","Depreciation","DepreciationDepreciation and provisions","Depreciation and amortization"],
    "(AOP222)Other operating expenses": ["Other operating expenses","Other expenses","Intangible costs","Other expenses and losses"],
    "(AOP223)Income from financial transactions": ["Income from financial transactions (financial income)","III. FINANCIAL INCOME","Financial income"],
    "(AOP234)Financial costs": ["Financial costs"],
    "(AOP250+/AOP251-)Profit or loss before taxation": ["Profit or loss before taxation","Profit before taxation","Loss before taxation","Profit before taxation/Loss before taxation","Profit Before Tax"],
    "(AOP252)Profit tax": ["Profit tax","Income tax","Taxes,duties and similar expenses","Taxes","Tax","Tax charge","Tax expense of the period"],
    "(AOP255+/AOP256-)Profit or loss after taxation": ["Profit after taxation/Loss after taxation"],
}
allowed_fields_slovenia = {
    "(AOP001)Fixed assets": ["Fixed assets"],
    "(AOP002)Intangible assets": ["I. Intangible assets", "Intangible assets", "Intangible fixed assets", "Total Intangible assets"],
    "(AOP004)Concessions,patents,licenses and similar rights and other intangible assets":["Concessions,industrial rights,licences","Concessions, patents, licenses, trade marks etc.", "Service marks,software and other intangible assets","Concessions,patents,licenses and similar rights and other intangible assets","Concessions, patents, licenses, trademarks, service marks, software and other intangible assets"],
    "(AOP005)Goodwill":["Goodwill"],
    "(AOP006)Advances for intangible assets and intangible assets in preparation":["Advances for intangible assets and intangible assets in preparation"],
    "(AOP008)Other intangible fixed assets":["Other intangible fixed assets","Other Intangible Assets"],
    "(AOP009)Tangible fixed assets": ["II. Tangible assets","Tangible fixed assets", "Total tangible assets" , "Property,plants,equipment and biological assets", "Immovables,plants and equipment","Immovables, plants and equipment","Tangible assets","Property, plants, equipment and biological assets"],
    "(AOP010)Land and buildings":["Land and buildings","Property, plant and equipment","Land & Buildings"],
    "(AOP011)Land":["Land","Land and buildings"],
    "(AOP012)Buildings":["Buildings","Land and buildings"],
    "(AOP013)Machinery and equipment": ["Machinery and equipment", "Plant and Machinery", "Technical equipment and machinery", "Property,plant and equipment","Plant and equipment","Plant & Machinery"],
    "(AOP014)Other equipment, furniture, fittings, tools, fixtures, vehicles": ["Other equipment, furniture, fittings, tools, fixtures, vehicles"],
    "(AOP016)Biological property":["Biological property","Biological assets"],
    "(AOP017)Advance payments for tangible assets": ["Advance payments for tangible assets", "Advances in property,plant,equipment and biological assets and property,biological asset in preparation","Advances in property, plant, equipment and biological assets and property, plant, equipment and biological assets in preparation"],
    "(AOP018)Tangible assets in progress": ["Tangible assets in progress", "Payments and fixed goods under construction","Immovables, plant  and equipment under construction"],
    "(AOP019)Other tangible assets": ["Other tangible assets","Other tangible fixed assets", "Other equipment" ,"Other Immovables,plants and equipment","Other immovables, plant and equipment and investment in third-party immovables, plant and equipment","Other equipment, furniture, fittings, tools, fixtures, vehicles","Other unspecified material fixed assets"],
    "(AOP020)Investments in real estate": ["Investment property measured at the cost model"],
    "(AOP021)Financial fixed assets": ["Financial fixed assets", "Long term investments","Long term financial investments and long term receivables Financial fixed assets","LONG-TERM FINANCIAL INVESTMENTS AND LONG-TERM RECEIVABLES","Financial assets"],
    "(AOP024)Loans to Group":["Loans to Group","Long-term loans to parent and subsidiary legal entities"],
    "(AOP025)Long term loans":["Long term loans","Long-term loans to legal entities with equity participation (excluding subsidiaries)"],
    "(AOP031)Long-term receivables": ["Long-term receivables", "Other long term investments and receivables","Long-term operating receivables"],
    "(AOP035)Deferred tax assets": ["DEFERRED TAX ASSETS", "Deferred tax assets"],
    "(AOP030)Other financial fixed assets":["Other financial fixed assets"],
    "(AOP034)Other long term receivables":["Other long term receivables","Other long-term receivables"],
    "(AOP036)Short-term assets": ["Short term assets", "Current assets", "Total Current assets", "Short term assets"],
    "(AOP037)Inventory": ["Inventories"],
    "(AOP038)Raw materials, consumabeles and supplies":["Raw materials, consumabeles and supplies", "Inventory-raw materials, consumabeles and supplies","Inventory - raw materials, consumables, working inventory", "Inventory of materials(fabrication material, spare parts, small inventory and car tires)","Inventory of materials (fabrication material, spare parts, small inventory and car tires"],
    "(AOP039)Inventory of materials":["Inventory of materials(fabricationmaterial, spare parts)"],
    "(AOP040)Work in progress":["Work in progress"],
    "(AOP041)Finished products and goods":["Finished products and goods", "Finished products/merchandise", "Finished goods", "Goods"],
    "(AOP042)Traiding Goods":["Traiding Goods","Inventory-Trading Goods","Inventory - Trading Goods"],
    "(AOP048)Prepayments":["Prepayments","Payments in advance for stock"],
    "(AOP045)Short-term receivables": ["Short-term operating receivables"],
    "(AOP046)Short-term intercompany receivables": ["Short-term intercompany receivables", "Receivables from parent companies and subsidiaries", "Receivables from affiliated companies", "Intercompany Receivables","Receivables from foreign parent companies, subsidiaries and other associated companies","Group Receivables"],
    "(AOP047)Short-term trade receivables": ["Short-term trade receivables", "Receivables from buyers", "Trade debtor" , "Trade receivables","Trade debtors"],
    "(AOP050)Short-term receivables from employees": ["Short-term receivables from employees"],
    "(AOP049)Receivables from the state and other institutions": ["Prepaid corporate income tax","Receivables from the state and other institutions"],
    "(AOP051)Other short-term receivables": ["Other short-term receivables","Other short term receivables", "Other receivables", "Miscellaneous receivables"],
    "(AOP052)Short-term financial assets": ["SHORT TERM FINANCIAL ASSETS","short-term financial assets", "Financial investment", "Short term financial assets","Short-term financial investments","Short-term investments"],
    "(AOP058)Other short-term financial investments":["other short-term financial investments","Other short term financial assets"],
    "(AOP059_060)Cash": ["Cash","Cash and cash equivalents", "Cash assets", "Bank balance cheques and cash on hand", "Cash and cash equivalent", "Cash","Bank balance, cheques and cash on hand","Liquid assets"],
    "(AOP062)Prepaid expenses": ["DEFERRED COSTS AND ACCRUED REVENUES"],
    "(AOP063)Total assets": ["TOTAL ASSETS", "Total assets", "Balance sheet total"],
    "(AOP064)Off balance sheet items": ["Off balance sheet items"],
    "(AOP065)Equity capital": ["Equity capital", "Owners equity"],
    "(AOP066)Subscribed and paid capital": ["Called capital"],
    "(AOP067)Emission premium":["Emission premium","Share premium"],
    "(AOP068)Own shares":["Own shares","Called up share capital","Called up share capital (issued capital stock)","Share capital","Subscribed capital unpaid"],
    "(AOP069)Subscribed unpaid capital":["Subscribed unpaid capital","Unpaid issued capital","Subscribed capital unpaid"],
    "(AOP071)Capital reserves": ["CAPITAL RESERVES", "Capital reserves", "Reserves", "Revenue","Revenue Reserves","Capital reserve"],
    "(AOP070)Revaluation reserves": ["Revaluation reserves","POSITIVE REVALUATION RESERVES AND UNREALIZED PROFIT FROM FINANCIAL ASSETS AND OTHER ELEMENTS OF OTHER COMPREHENSIVE INCOME"],
    "(AOP072)Legal reserves":["Legal reserves"],
    "(AOP074)Other reserves":["Other reserves"],
    "(AOP075+/AOP076-)Profit or loss carried forward": ["Profit or loss carried forward","Retained profit (earnings) of the year (net profits)","Accumulated retained earnings from previous periods","Retained profit", "Retained earnings brought forward", "Net retained profit/net acumalted loss", "Net profit or loss from previos years","Equity, net retained profits/net accumulated losses (balance sheet)","Net profit or loss from previous periods","Accumulated profit reserve","Retained earnings or loss"],
    "(AOP077+/AOP078-)Net profit or loss for the year": ["Net profit or loss for the year","Current period profit", "Loss from previos years", "Net profit for the period" , "Profit of the year","Net profit or loss for the year", "Net profit for the period","Retained earnings for the current year"],
    "(AOP081)Liabilities": ["Liabilities", "Total liabilities","Total debt","Total debts"],
    "(AOP082)Provision for risks and charges":["Provision for risks and charges","Provisions for risks and charges"],
    "(AOP083)Provisions for pensions and similar obligations":["Provisions for pensions and similar obligations"],
    "(AOP084)Other provisions":["Other provisions"],
    "(AOP085)Long-term liabilities": ["Long-term liabilities", "Long term liabilities", "Total Long term liabilities","Long-term financial and operating liabilities","Long-term provisions and long-term liabilities"],
    "(AOP086)Long-term liabilities to affiliates": ["Long-term liabilities to affiliates", "Group payables due after 1 year", "Long term liabilities to affiliates"],
    "(AOP087)Trade liabilities":["Trade liabilities"],
    "(AOP090)Long-term liabilities for loans": ["Long-term financial liabilities"],
    "(AOP092)Other loans/finance due after 1 year":["Long-term operating liabilities"],
    "(AOP093)Other long-term liabilities": ["Other long-term liabilities", "Other long term liabilities"],
    "(AOP095)Short-term liabilities": ["Short-term financial and operating liabilities"],
    "(AOP096)Short-term liabilities to affiliates": ["Liabilities of disposal groups"],
    "(AOP098)Prepayments, deposits and gurantee":["Prepayments, deposits and gurantee","Liabilities for securities","Prepayments, deposits and guarantees"],
    "(AOP103)Liabilities for loans, deposits, etc. to companies within the group": ["Liabilities for loans, deposits, etc. to companies within the group"],
    "(AOP097)Short-term trade creditors": ["Short-term trade creditors", "Liabilities to suppliers","Trade Payables","Trade creditors(acounts payable)", " Trade payables","Commitments towards suppliers","Trade payable"],
    "(AOP100)Short-term liabilities to employees": ["Short-term liabilities to employees","Liabilities towards employees"],
    "(AOP099)Short-term liabilities for taxes, contributions and other fees": ["Short-term liabilities for taxes, contributions and other fees"],
    "(AOP101)Liability for income tax":["Liability for income tax","S-term Liabilities for tax,contributions and other fees"],
    "(AOP104)Obligations for taken loans and credits":["Obligations for taken loans and credits","Short Term Bank Debt","Short liabilities for loans","Short term loans","Short term financial liabilities","Obligation for taken loans","BANK LOANS AND OVERDRAFT","Liabilities to bank","Loans liabilities from credit institutions","Short-term liabilities for loans","Short-term financial liabilities","S-term Liabilities payable to financial institutions","Banks loans and overdraft","Other Loans/Finance"],
    "(AOP108)Other short-term liabilities": ["Other short term liabilities", "Other operating and liabilities and other short term liabilities", "Other short-term liabilities ",  "Other  liabilities  including accruals",  "Other  liabilities","Other short-term liabilities","Miscellaneous Liabilities"],
    "(AOP109)Accruals and deferred income": ["Accruals and deferred income"],
    "(AOP111)Total liabilities and funds": [""],
    "(AOP112)Off balance sheet items": ["Off balance sheet items"],
    "(AOP201)Turnover, sales revenue": ["Turnover"],
    "(AOP202)Revenues from sales": ["Operating income"],
    "(AOP203)Other income(other revenues)": ["Other income(other revenues)","Other operating income (outside the group)","Other operating income","OTHER OPERATING REVENUE","Other income and profits","Income from financial transactions (financial income)","Other income"],
    "(AOP206)Own work capitalized": ["Own work capitalized"],
    "(AOP207)Operating expenses": ["TOTAL EXPENSES"],
    "(AOP208)Material costs": ["Material costs","Cost of raw materials and consumables","RAW MATERIAL COSTS, FUEL AND ENERGY COSTS"],
    "(AOP209)Cost of goods sold": ["Costs of goods, material and services"],
    "(AOP213)Staff costs": ["Labour costs"],
    "(AOP218)Depreciation on fixed assets": ["Depreciation and amortization costs"],
    "(AOP222)Other operating expenses": ["Other operating expenses"],
    "(AOP223)Income from financial transactions": ["Financial revenues"],
    "(AOP234)Financial costs": ["Financial expenses"],
    "(AOP250+/AOP251-)Profit or loss before taxation": ["Profit or loss before taxation","Profit before taxation","Loss before taxation","Profit before taxation/Loss before taxation","Profit Before Tax"],
    "(AOP252)Profit tax": ["Profit tax","Income tax","Taxes,duties and similar expenses","Taxes","Tax","Tax charge","Tax expense of the period"],
    "(AOP255+/AOP256-)Profit or loss after taxation": ["Net profit for the period"],
}
allowed_fields_montenegro = {
    "(AOP001)Fixed assets": ["Fixed assets"],
    "(AOP002)Intangible assets": ["I. Intangible assets", "Intangible assets", "Intangible fixed assets", "Total Intangible assets"],
    "(AOP004)Concessions,patents,licenses and similar rights and other intangible assets":["Concessions, patents, licenses and similar rights and other intangible assets"],
    "(AOP005)Goodwill":["Goodwill"],
    "(AOP006)Advances for intangible assets and intangible assets in preparation":["Advances for intangible assets and intangible assets in preparation"],
    "(AOP008)Other intangible fixed assets":["Other intangible fixed assets","Other Intangible Assets"],
    "(AOP009)Tangible fixed assets": ["II. Tangible assets","Tangible fixed assets", "Total tangible assets" , "Property,plants,equipment and biological assets", "Immovables,plants and equipment","Immovables, plants and equipment","Tangible assets","Property, plants, equipment and biological assets"],
    "(AOP010)Land and buildings":[""],
    "(AOP011)Land":["Land and buildings"],
    "(AOP012)Buildings":[""],
    "(AOP013)Machinery and equipment": ["Machinery and equipment", "Plant and Machinery", "Technical equipment and machinery", "Property,plant and equipment","Plant and equipment","Plant & Machinery"],
    "(AOP014)Other equipment, furniture, fittings, tools, fixtures, vehicles": ["Other equipment, furniture, fittings, tools, fixtures, vehicles"],
    "(AOP016)Biological property":["Biological property","Biological assets"],
    "(AOP017)Advance payments for tangible assets": ["Advance payments for tangible assets", "Advances in property,plant,equipment and biological assets and property,biological asset in preparation","Advances in property, plant, equipment and biological assets and property, plant, equipment and biological assets in preparation"],
    "(AOP018)Tangible assets in progress": ["Tangible assets in progress", "Payments and fixed goods under construction","Immovables, plant  and equipment under construction"],
    "(AOP019)Other tangible assets": ["Other tangible assets","Other tangible fixed assets", "Other equipment" ,"Other Immovables,plants and equipment","Other immovables, plant and equipment and investment in third-party immovables, plant and equipment","Other equipment, furniture, fittings, tools, fixtures, vehicles","Other unspecified material fixed assets"],
    "(AOP020)Investments in real estate": ["Investments in real estate","Investment property measured at the cost model"],
    "(AOP021)Financial fixed assets": ["Financial fixed assets", "Long term investments","Long term financial investments and long term receivables Financial fixed assets","LONG-TERM FINANCIAL INVESTMENTS AND LONG-TERM RECEIVABLES","Financial assets"],
    "(AOP024)Loans to Group":[""],
    "(AOP025)Long term loans":["Long term loans","Long-term loans to legal entities with equity participation (excluding subsidiaries)"],
    "(AOP031)Long-term receivables": ["Long-term receivables", "Other long term investments and receivables","Long-term operating receivables"],
    "(AOP035)Deferred tax assets": ["DEFERRED TAX ASSETS", "Deferred tax assets"],
    "(AOP030)Other financial fixed assets":["Other long-term investments and receivables"],
    "(AOP034)Other long term receivables":["Other long term receivables","Other long-term receivables"],
    "(AOP036)Short-term assets": ["Short term assets", "Current assets", "Total Current assets", "Short term assets"],
    "(AOP037)Inventory": ["Inventories"],
    "(AOP038)Raw materials, consumabeles and supplies":["Inventory of materials (fabrication material, spare parts, small inventory and car tires)"],
    "(AOP039)Inventory of materials":["Inventory of materials(fabricationmaterial, spare parts)"],
    "(AOP040)Work in progress":["Work in progress"],
    "(AOP041)Finished products and goods":["Finished products and goods", "Finished products/merchandise", "Finished goods", "Goods"],
    "(AOP042)Traiding Goods":["Traiding Goods","Inventory-Trading Goods","Inventory - Trading Goods"],
    "(AOP048)Prepayments":["Prepayments","Payments in advance for stock"],
    "(AOP045)Short-term receivables": ["Short-term receivables"],
    "(AOP046)Short-term intercompany receivables": ["Short-term intercompany receivables", "Receivables from parent companies and subsidiaries", "Receivables from affiliated companies", "Intercompany Receivables","Receivables from foreign parent companies, subsidiaries and other associated companies","Group Receivables"],
    "(AOP047)Short-term trade receivables": ["Short-term trade receivables", "Receivables from buyers", "Trade debtor" , "Trade receivables","Trade debtors"],
    "(AOP050)Short-term receivables from employees": ["Short-term receivables from employees"],
    "(AOP049)Receivables from the state and other institutions": ["Prepaid corporate income tax","Receivables from the state and other institutions"],
    "(AOP051)Other short-term receivables": ["Other short-term receivables","Other short term receivables", "Other receivables", "Miscellaneous receivables"],
    "(AOP052)Short-term financial assets": ["SHORT TERM FINANCIAL ASSETS","short-term financial assets", "Financial investment", "Short term financial assets","Short-term financial investments","Short-term investments"],
    "(AOP058)Other short-term financial investments":["other short-term financial investments","Other short term financial assets"],
    "(AOP059_060)Cash": ["Cash on accounts and in hand"],
    "(AOP062)Prepaid expenses": ["Prepaid expenses","ACCRUALS"],
    "(AOP063)Total assets": ["TOTAL ASSETS", "Total assets", "Balance sheet total"],
    "(AOP064)Off balance sheet items": ["Off balance sheet items"],
    "(AOP065)Equity capital": ["Equity capital", "Owners equity"],
    "(AOP066)Subscribed and paid capital": ["Subscribed and paid capital", "Basic capital", "Shareholders equity","Called capital", "Share capital", "Called and share capital","Subscribed capital","Called up share capital (issued capital stock)"],
    "(AOP067)Emission premium":["Emission premium","Share premium"],
    "(AOP068)Own shares":["Own shares","Called up share capital","Called up share capital (issued capital stock)","Share capital","Subscribed capital unpaid"],
    "(AOP069)Subscribed unpaid capital":[""],
    "(AOP071)Capital reserves": ["CAPITAL RESERVES", "Capital reserves", "Reserves", "Revenue","Revenue Reserves","Capital reserve"],
    "(AOP070)Revaluation reserves": ["Revaluation reserves","POSITIVE REVALUATION RESERVES AND UNREALIZED PROFIT FROM FINANCIAL ASSETS AND OTHER ELEMENTS OF OTHER COMPREHENSIVE INCOME"],
    "(AOP072)Legal reserves":["Legal reserves"],
    "(AOP074)Other reserves":["Other reserves"],
    "(AOP075+/AOP076-)Profit or loss carried forward": ["Profit or loss carried forward","Retained profit (earnings) of the year (net profits)","Accumulated retained earnings from previous periods","Retained profit", "Retained earnings brought forward", "Net retained profit/net acumalted loss", "Net profit or loss from previos years","Equity, net retained profits/net accumulated losses (balance sheet)","Net profit or loss from previous periods","Accumulated profit reserve","Retained earnings or loss"],
    "(AOP077+/AOP078-)Net profit or loss for the year": ["Net profit or loss for the year","Current period profit", "Loss from previos years", "Net profit for the period" , "Profit of the year","Net profit or loss for the year", "Net profit for the period","Retained earnings for the current year"],
    "(AOP081)Liabilities": ["Liabilities", "Total liabilities","Total debt","Total debts"],
    "(AOP082)Provision for risks and charges":["Provision for risks and charges","Provisions for risks and charges"],
    "(AOP083)Provisions for pensions and similar obligations":["Provisions for pensions and similar obligations"],
    "(AOP084)Other provisions":["Other provisions"],
    "(AOP085)Long-term liabilities": ["Long-term liabilities", "Long term liabilities", "Total Long term liabilities","Long-term financial and operating liabilities","Long-term provisions and long-term liabilities"],
    "(AOP086)Long-term liabilities to affiliates": ["Long-term liabilities to affiliates", "Group payables due after 1 year", "Long term liabilities to affiliates"],
    "(AOP087)Trade liabilities":["Trade liabilities"],
    "(AOP090)Long-term liabilities for loans": ["Long-term loans"],
    "(AOP092)Other loans/finance due after 1 year":["Other loans/finance due after 1 year","Long term liabilities payable to financial institutions"],
    "(AOP093)Other long-term liabilities": ["Other long-term liabilities", "Other long term liabilities"],
    "(AOP095)Short-term liabilities": ["Short-term liabilities","IV. SHORT-TERM LIABILITIES","Short- term liabilities","Short term liabilities","SHORT- TERM LIABILITIES", "TOTAL CURRENT LIABILITIES", "Total current liabilities","Short- term liabilities and short term provisions"," Short- term financial and operating liabilities"],
    "(AOP096)Short-term liabilities to affiliates": ["Short-term liabilities to affiliates","Liabilities of disposal groups", " Trade payables-foreing parent company, subsidiares and other associated companies", " Liabilities to affiliated companies","Group payables"],
    "(AOP098)Prepayments, deposits and gurantee":["Prepayments, deposits and gurantee","Liabilities for securities","Prepayments, deposits and guarantees"],
    "(AOP103)Liabilities for loans, deposits, etc. to companies within the group": ["Liabilities for loans, deposits, etc. to companies within the group"],
    "(AOP097)Short-term trade creditors": ["Short-term trade creditors", "Liabilities to suppliers","Trade Payables","Trade creditors(acounts payable)", " Trade payables","Commitments towards suppliers","Trade payable"],
    "(AOP100)Short-term liabilities to employees": ["Short-term liabilities to employees","Liabilities towards employees"],
    "(AOP099)Short-term liabilities for taxes, contributions and other fees": ["Short-term liabilities for taxes, contributions and other fees"],
    "(AOP101)Liability for income tax":["Liabilities for value added tax and other public revenues"],
    "(AOP104)Obligations for taken loans and credits":["Obligations for taken loans and credits","Short Term Bank Debt","Short liabilities for loans","Short term loans","Short term financial liabilities","Obligation for taken loans","BANK LOANS AND OVERDRAFT","Liabilities to bank","Loans liabilities from credit institutions","Short-term liabilities for loans","Short-term financial liabilities","S-term Liabilities payable to financial institutions","Banks loans and overdraft","Other Loans/Finance"],
    "(AOP108)Other short-term liabilities": ["Other short term liabilities", "Other operating and liabilities and other short term liabilities", "Other short-term liabilities ",  "Other  liabilities  including accruals",  "Other  liabilities","Other short-term liabilities","Miscellaneous Liabilities"],
    "(AOP109)Accruals and deferred income": ["Accruals and deferred income"],
    "(AOP111)Total liabilities and funds": ["Total liabilities and funds", "TOTAL EQUITY AND LIABILITIES "," TOTAL LIABILITIES","BALANCE SHEET TOTAL","TOTAL LIABILITIES","Balance sheet total"],
    "(AOP112)Off balance sheet items": ["Off balance sheet items"],
    "(AOP201)Turnover, sales revenue": ["Total operating income"],
    "(AOP202)Revenues from sales": ["Turnover, sales revenue"],
    "(AOP203)Other income(other revenues)": ["Other income(other revenues)","Other operating income (outside the group)","Other operating income","OTHER OPERATING REVENUE","Other income and profits","Income from financial transactions (financial income)","Other income"],
    "(AOP206)Own work capitalized": ["Own work capitalized"],
    "(AOP207)Operating expenses": ["Operating expenses","Operating Costs"],
    "(AOP208)Material costs": ["Material costs","Cost of raw materials and consumables","RAW MATERIAL COSTS, FUEL AND ENERGY COSTS"],
    "(AOP209)Cost of goods sold": ["Cost of goods sold","The purchase value of the goods sold","Cost of materials (type of expenditure format)","Cost of goods sold and the cost of materials"],
    "(AOP213)Staff costs": ["Staff costs (employee costs)","Wages and salaries","Wages & Salaries","Short-term financial and operating liabilities","Personnel type expenses","Wages expenses, wage compensation and other personal expenses"],
    "(AOP218)Depreciation on fixed assets": ["Depreciation on fixed assets","Depreciation","DepreciationDepreciation and provisions","Depreciation and amortization"],
    "(AOP222)Other operating expenses": ["Other operating expenses","Other expenses","Intangible costs","Other expenses and losses"],
    "(AOP223)Income from financial transactions": ["Income from financial transactions (financial income)","III. FINANCIAL INCOME","Financial income"],
    "(AOP234)Financial costs": ["Financial costs"],
    "(AOP250+/AOP251-)Profit or loss before taxation": ["Profit or loss before taxation","Profit before taxation","Loss before taxation","Profit before taxation/Loss before taxation","Profit Before Tax"],
    "(AOP252)Profit tax": ["Profit tax","Income tax","Taxes,duties and similar expenses","Taxes","Tax","Tax charge","Tax expense of the period"],
    "(AOP255+/AOP256-)Profit or loss after taxation": ["Profit or loss after taxation","Profit after taxation","Loss after taxation","TOTAL RESULT"],
}
allowed_fields_romania = {
    "(AOP001)Fixed assets": ["Fixed assets"],
    "(AOP002)Intangible assets": ["I. Intangible assets", "Intangible assets", "Intangible fixed assets", "Total Intangible assets"],
    "(AOP004)Concessions,patents,licenses and similar rights and other intangible assets":["Concessions,industrial rights,licences","Concessions, patents, licenses, trade marks etc.", "Service marks,software and other intangible assets","Concessions,patents,licenses and similar rights and other intangible assets","Concessions, patents, licenses, trademarks, service marks, software and other intangible assets"],
    "(AOP005)Goodwill":["Goodwill"],
    "(AOP006)Advances for intangible assets and intangible assets in preparation":["Advances for intangible assets and intangible assets in preparation"],
    "(AOP008)Other intangible fixed assets":["Other intangible fixed assets","Other Intangible Assets"],
    "(AOP009)Tangible fixed assets": ["II. Tangible assets","Tangible fixed assets", "Total tangible assets" , "Property,plants,equipment and biological assets", "Immovables,plants and equipment","Immovables, plants and equipment","Tangible assets","Property, plants, equipment and biological assets"],
    "(AOP010)Land and buildings":["Land and buildings","Property, plant and equipment","Land & Buildings"],
    "(AOP011)Land":["Land","Land and buildings"],
    "(AOP012)Buildings":["Buildings","Land and buildings"],
    "(AOP013)Machinery and equipment": ["Machinery and equipment", "Plant and Machinery", "Technical equipment and machinery", "Property,plant and equipment","Plant and equipment","Plant & Machinery"],
    "(AOP014)Other equipment, furniture, fittings, tools, fixtures, vehicles": ["Other equipment, furniture, fittings, tools, fixtures, vehicles"],
    "(AOP016)Biological property":["Biological property","Biological assets"],
    "(AOP017)Advance payments for tangible assets": ["Advance payments for tangible assets", "Advances in property,plant,equipment and biological assets and property,biological asset in preparation","Advances in property, plant, equipment and biological assets and property, plant, equipment and biological assets in preparation"],
    "(AOP018)Tangible assets in progress": ["Tangible assets in progress", "Payments and fixed goods under construction","Immovables, plant  and equipment under construction"],
    "(AOP019)Other tangible assets": ["Other tangible assets","Other tangible fixed assets", "Other equipment" ,"Other Immovables,plants and equipment","Other immovables, plant and equipment and investment in third-party immovables, plant and equipment","Other equipment, furniture, fittings, tools, fixtures, vehicles","Other unspecified material fixed assets"],
    "(AOP020)Investments in real estate": ["Investments in real estate","Investment property measured at the cost model"],
    "(AOP021)Financial fixed assets": ["Financial fixed assets", "Long term investments","Long term financial investments and long term receivables Financial fixed assets","LONG-TERM FINANCIAL INVESTMENTS AND LONG-TERM RECEIVABLES","Financial assets"],
    "(AOP024)Loans to Group":["Loans to Group","Long-term loans to parent and subsidiary legal entities"],
    "(AOP025)Long term loans":["Long term loans","Long-term loans to legal entities with equity participation (excluding subsidiaries)"],
    "(AOP031)Long-term receivables": ["Long-term receivables", "Other long term investments and receivables","Long-term operating receivables"],
    "(AOP035)Deferred tax assets": ["DEFERRED TAX ASSETS", "Deferred tax assets"],
    "(AOP030)Other financial fixed assets":["Other financial fixed assets"],
    "(AOP034)Other long term receivables":["Other long term receivables","Other long-term receivables"],
    "(AOP036)Short-term assets": ["Short term assets", "Current assets", "Total Current assets", "Short term assets"],
    "(AOP037)Inventory": ["Inventory", "Total inventories", "Inventories"],
    "(AOP038)Raw materials, consumabeles and supplies":["Raw materials, consumabeles and supplies", "Inventory-raw materials, consumabeles and supplies","Inventory - raw materials, consumables, working inventory", "Inventory of materials(fabrication material, spare parts, small inventory and car tires)","Inventory of materials (fabrication material, spare parts, small inventory and car tires"],
    "(AOP039)Inventory of materials":["Inventory of materials(fabricationmaterial, spare parts)"],
    "(AOP040)Work in progress":["Work in progress"],
    "(AOP041)Finished products and goods":["Finished products and goods", "Finished products/merchandise", "Finished goods", "Goods"],
    "(AOP042)Traiding Goods":["Traiding Goods","Inventory-Trading Goods","Inventory - Trading Goods"],
    "(AOP048)Prepayments":["Prepayments","Payments in advance for stock"],
    "(AOP045)Short-term receivables": ["Short-term receivables"],
    "(AOP046)Short-term intercompany receivables": ["Short-term intercompany receivables", "Receivables from parent companies and subsidiaries", "Receivables from affiliated companies", "Intercompany Receivables","Receivables from foreign parent companies, subsidiaries and other associated companies","Group Receivables"],
    "(AOP047)Short-term trade receivables": ["Short-term trade receivables", "Receivables from buyers", "Trade debtor" , "Trade receivables","Trade debtors"],
    "(AOP050)Short-term receivables from employees": ["Short-term receivables from employees"],
    "(AOP049)Receivables from the state and other institutions": ["Prepaid corporate income tax","Receivables from the state and other institutions"],
    "(AOP051)Other short-term receivables": ["Other short-term receivables","Other short term receivables", "Other receivables", "Miscellaneous receivables"],
    "(AOP052)Short-term financial assets": ["SHORT TERM FINANCIAL ASSETS","short-term financial assets", "Financial investment", "Short term financial assets","Short-term financial investments","Short-term investments"],
    "(AOP058)Other short-term financial investments":["other short-term financial investments","Other short term financial assets"],
    "(AOP059_060)Cash": ["Cash","Cash and cash equivalents", "Cash assets", "Bank balance cheques and cash on hand", "Cash and cash equivalent", "Cash","Bank balance, cheques and cash on hand","Liquid assets"],
    "(AOP062)Prepaid expenses": ["Prepaid expenses","ACCRUALS"],
    "(AOP063)Total assets": ["TOTAL ASSETS", "Total assets", "Balance sheet total"],
    "(AOP064)Off balance sheet items": ["Off balance sheet items"],
    "(AOP065)Equity capital": ["Equity capital", "Owners equity"],
    "(AOP066)Subscribed and paid capital": ["Subscribed and paid capital", "Basic capital", "Shareholders equity","Called capital", "Share capital", "Called and share capital","Subscribed capital","Called up share capital (issued capital stock)"],
    "(AOP067)Emission premium":["Emission premium","Share premium"],
    "(AOP068)Own shares":["Own shares","Called up share capital","Called up share capital (issued capital stock)","Share capital","Subscribed capital unpaid"],
    "(AOP069)Subscribed unpaid capital":["Subscribed unpaid capital","Unpaid issued capital","Subscribed capital unpaid"],
    "(AOP071)Capital reserves": ["CAPITAL RESERVES", "Capital reserves", "Reserves", "Revenue","Revenue Reserves","Capital reserve"],
    "(AOP070)Revaluation reserves": ["Revaluation reserves","POSITIVE REVALUATION RESERVES AND UNREALIZED PROFIT FROM FINANCIAL ASSETS AND OTHER ELEMENTS OF OTHER COMPREHENSIVE INCOME"],
    "(AOP072)Legal reserves":["Legal reserves"],
    "(AOP074)Other reserves":["Other reserves"],
    "(AOP075+/AOP076-)Profit or loss carried forward": ["Profit or loss carried forward","Retained profit (earnings) of the year (net profits)","Accumulated retained earnings from previous periods","Retained profit", "Retained earnings brought forward", "Net retained profit/net acumalted loss", "Net profit or loss from previos years","Equity, net retained profits/net accumulated losses (balance sheet)","Net profit or loss from previous periods","Accumulated profit reserve","Retained earnings or loss"],
    "(AOP077+/AOP078-)Net profit or loss for the year": ["Net profit or loss for the year","Current period profit", "Loss from previos years", "Net profit for the period" , "Profit of the year","Net profit or loss for the year", "Net profit for the period","Retained earnings for the current year"],
    "(AOP081)Liabilities": ["Liabilities", "Total liabilities","Total debt","Total debts"],
    "(AOP082)Provision for risks and charges":["Provision for risks and charges","Provisions for risks and charges"],
    "(AOP083)Provisions for pensions and similar obligations":["Provisions for pensions and similar obligations"],
    "(AOP084)Other provisions":["Other provisions"],
    "(AOP085)Long-term liabilities": ["Long-term liabilities", "Long term liabilities", "Total Long term liabilities","Long-term financial and operating liabilities","Long-term provisions and long-term liabilities"],
    "(AOP086)Long-term liabilities to affiliates": ["Long-term liabilities to affiliates", "Group payables due after 1 year", "Long term liabilities to affiliates"],
    "(AOP087)Trade liabilities":["Trade liabilities"],
    "(AOP090)Long-term liabilities for loans": ["Long-term liabilities for loans"],
    "(AOP092)Other loans/finance due after 1 year":["Other loans/finance due after 1 year","Long term liabilities payable to financial institutions"],
    "(AOP093)Other long-term liabilities": ["Other long-term liabilities", "Other long term liabilities"],
    "(AOP095)Short-term liabilities": ["Short-term liabilities","IV. SHORT-TERM LIABILITIES","Short- term liabilities","Short term liabilities","SHORT- TERM LIABILITIES", "TOTAL CURRENT LIABILITIES", "Total current liabilities","Short- term liabilities and short term provisions"," Short- term financial and operating liabilities"],
    "(AOP096)Short-term liabilities to affiliates": ["Short-term liabilities to affiliates","Liabilities of disposal groups", " Trade payables-foreing parent company, subsidiares and other associated companies", " Liabilities to affiliated companies","Group payables"],
    "(AOP098)Prepayments, deposits and gurantee":["Prepayments, deposits and gurantee","Liabilities for securities","Prepayments, deposits and guarantees"],
    "(AOP103)Liabilities for loans, deposits, etc. to companies within the group": ["Liabilities for loans, deposits, etc. to companies within the group"],
    "(AOP097)Short-term trade creditors": ["Short-term trade creditors", "Liabilities to suppliers","Trade Payables","Trade creditors(acounts payable)", " Trade payables","Commitments towards suppliers","Trade payable"],
    "(AOP100)Short-term liabilities to employees": ["Short-term liabilities to employees","Liabilities towards employees"],
    "(AOP099)Short-term liabilities for taxes, contributions and other fees": ["Short-term liabilities for taxes, contributions and other fees"],
    "(AOP101)Liability for income tax":["Liability for income tax","S-term Liabilities for tax,contributions and other fees"],
    "(AOP104)Obligations for taken loans and credits":["Obligations for taken loans and credits","Short Term Bank Debt","Short liabilities for loans","Short term loans","Short term financial liabilities","Obligation for taken loans","BANK LOANS AND OVERDRAFT","Liabilities to bank","Loans liabilities from credit institutions","Short-term liabilities for loans","Short-term financial liabilities","S-term Liabilities payable to financial institutions","Banks loans and overdraft","Other Loans/Finance"],
    "(AOP108)Other short-term liabilities": ["Other short term liabilities", "Other operating and liabilities and other short term liabilities", "Other short-term liabilities ",  "Other  liabilities  including accruals",  "Other  liabilities","Other short-term liabilities","Miscellaneous Liabilities"],
    "(AOP109)Accruals and deferred income": ["Accruals and deferred income"],
    "(AOP111)Total liabilities and funds": ["Total liabilities and funds", "TOTAL EQUITY AND LIABILITIES "," TOTAL LIABILITIES","BALANCE SHEET TOTAL","TOTAL LIABILITIES","Balance sheet total"],
    "(AOP112)Off balance sheet items": ["Off balance sheet items"],
    "(AOP201)Turnover, sales revenue": ["Turnover, sales revenue","Revenues from contracts with customers","Operating income","Turnover","Total income"],
    "(AOP202)Revenues from sales": ["Revenues from sales","Income from sales","Income from sales (outside group)","Income from goods sold","Income from sales of goods","Turnover","Operating income","Net sales revenue"],
    "(AOP203)Other income(other revenues)": ["Other income(other revenues)","Other operating income (outside the group)","Other operating income","OTHER OPERATING REVENUE","Other income and profits","Income from financial transactions (financial income)","Other income"],
    "(AOP206)Own work capitalized": ["Own work capitalized"],
    "(AOP207)Operating expenses": ["Operating expenses","Operating Costs"],
    "(AOP208)Material costs": ["Material costs","Cost of raw materials and consumables","RAW MATERIAL COSTS, FUEL AND ENERGY COSTS"],
    "(AOP209)Cost of goods sold": ["Cost of goods sold","The purchase value of the goods sold","Cost of materials (type of expenditure format)","Cost of goods sold and the cost of materials"],
    "(AOP213)Staff costs": ["Staff costs (employee costs)","Wages and salaries","Wages & Salaries","Short-term financial and operating liabilities","Personnel type expenses","Wages expenses, wage compensation and other personal expenses"],
    "(AOP218)Depreciation on fixed assets": ["Depreciation on fixed assets","Depreciation","DepreciationDepreciation and provisions","Depreciation and amortization"],
    "(AOP222)Other operating expenses": ["Other operating expenses","Other expenses","Intangible costs","Other expenses and losses"],
    "(AOP223)Income from financial transactions": ["Income from financial transactions (financial income)","III. FINANCIAL INCOME","Financial income"],
    "(AOP234)Financial costs": ["Financial costs"],
    "(AOP250+/AOP251-)Profit or loss before taxation": ["Profit or loss before taxation","Profit before taxation","Loss before taxation","Profit before taxation/Loss before taxation","Profit Before Tax"],
    "(AOP252)Profit tax": ["Profit tax","Income tax","Taxes,duties and similar expenses","Taxes","Tax","Tax charge","Tax expense of the period"],
    "(AOP255+/AOP256-)Profit or loss after taxation": ["Profit or loss after taxation","Profit after taxation","Loss after taxation","TOTAL RESULT"],
}



# Map countries to alias dictionaries
country_alias_map = {
    "Serbia": allowed_fields_serbia,
    "Netherlands": allowed_fields_netherlands,
    "France": allowed_fields_france,
    "Germany": allowed_fields_germany,
    "Croatia": allowed_fields_croatia,
    "Albania": allowed_fields_albania,
    "Belgium": allowed_fields_belgium,
    "Bulgaria": allowed_fields_bulgaria,
    "Slovenia": allowed_fields_slovenia,
    "Montenegro": allowed_fields_montenegro,
    "Romania": allowed_fields_romania
}

# ------------------------------
# Streamlit page config
# ------------------------------
st.set_page_config(page_title="Coface JSON", layout="wide")
st.markdown("""
<style>
#MainMenu {visibility: hidden;}
header {visibility: hidden;}
footer {visibility: hidden;}
</style>
""", unsafe_allow_html=True)

st.title("Convert COFACE JSON fields to Excel")

country = st.selectbox("Select country", list(country_alias_map.keys()))
uploaded_file = st.file_uploader("Upload JSON file", type=["json"])

# ------------------------------
# Helper functions
# ------------------------------
def convert_date(value):
    value = str(value)
    if value.endswith("0000"):
        return value[:4]
    if len(value) == 8:
        return f"{value[6:8]}.{value[4:6]}.{value[:4]}"
    return value

def format_amount_list(lst):
    formatted = []
    for a in lst:
        if a is None:
            continue
        num = float(a)
        formatted.append(
            f"{num:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        )
    return "; ".join(formatted)

def dedupe(seq):
    seen = set()
    out = []
    for x in seq:
        if x not in seen:
            seen.add(x)
            out.append(x)
    return out

def extract_aop_code(field_name):
    match = re.search(r"\((AOP\d+)\)", field_name)
    return match.group(1) if match else ""

# ------------------------------
# Recursive extraction
# ------------------------------
def extract_values(obj, extracted, allowed_fields, parent_dates=None, parent_amounts=None):
    if parent_dates is None:
        parent_dates = []
    if parent_amounts is None:
        parent_amounts = []

    if isinstance(obj, dict):

        current_dates = parent_dates.copy()
        current_amounts = parent_amounts.copy()

        # Capture date or fromAmount
        if "date" in obj:
            current_dates.append(obj["date"])
        if "fromAmount" in obj:
            current_amounts.append(obj["fromAmount"])

        name = None
        value = None

        # CASE 1: direct name/value
        if "name" in obj and "value" in obj:
            name = str(obj["name"]).strip()
            value = obj.get("value")

        # CASE 2: nested under "type"
        elif "type" in obj and isinstance(obj["type"], dict):
            t = obj["type"]
            if "name" in t and "value" in t:
                name = str(t["name"]).strip()
                value = t["value"]

        # If we captured valid name, map to field
        if name:
            for field, aliases in allowed_fields.items():
                for alias in aliases:
                    if name.lower() == alias.lower():
                        # Save value
                        if extracted[field]["value"] is None and value is not None:
                            extracted[field]["value"] = value
                        # Save matched alias name
                        extracted[field]["alias"] = alias
                        extracted[field]["date"].extend(current_dates)
                        extracted[field]["fromAmount"].extend(current_amounts)
                        break

        # Continue recursion
        for k, v in obj.items():
            extract_values(v, extracted, allowed_fields, current_dates, current_amounts)

    elif isinstance(obj, list):
        for item in obj:
            extract_values(item, extracted, allowed_fields, parent_dates, parent_amounts)

# ------------------------------
# Main processing
# ------------------------------
if uploaded_file:
    try:
        data = json.load(uploaded_file)
        allowed_fields = country_alias_map[country]

        # Init storage with alias field included
        extracted = {
            field: {"value": None, "alias": None, "date": [], "fromAmount": []}
            for field in allowed_fields
        }

        extract_values(data, extracted, allowed_fields)

        # Dedupe lists
        for field in extracted:
            extracted[field]["date"] = dedupe(extracted[field]["date"])
            extracted[field]["fromAmount"] = dedupe(extracted[field]["fromAmount"])

        # ---------------------------
        # Create DataFrame
        # ---------------------------
        df = pd.DataFrame(
            [
                (
                    field,
                    extracted[field]["value"] if extracted[field]["value"] is not None else "",
                    extracted[field]["alias"] if extracted[field]["alias"] else "",
                    f"{extract_aop_code(field)} {extracted[field]['value'] if extracted[field]['value'] is not None else ''}",
                    "; ".join(convert_date(d) for d in extracted[field]["date"]),
                    format_amount_list(extracted[field]["fromAmount"])
                )
                for field in allowed_fields
            ],
            columns=["Field", "Value", "NameValue", "NameMapping", "Date", "FromAmount"]
        )

        st.dataframe(df)

        # ---------------------------
        # Export to Excel
        # ---------------------------
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Sheet1")
            ws = writer.sheets["Sheet1"]

            # Autofit columns
            for col in ws.columns:
                max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 4

            # Bold first column
            for cell in ws["A"]:
                cell.font = Font(bold=True)

            # Add borders
            thin = Side(border_style="thin", color="000000")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = border

        output.seek(0)

        st.download_button(
            label="📥 Download Excel",
            data=output,
            file_name=f"mapped_values_{country}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        st.error(f"Error: {e}")




























































































































































